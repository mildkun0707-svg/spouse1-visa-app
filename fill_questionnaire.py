"""
配偶者ビザ 質問書 自動転記エンジン (全8ページ対応)
=====================================================
入力: form_data(dict) + case_flags(dict)  ※ master_schema.json のJSONそのまま
出力: 入管公式書式に転記された .xlsx

使い方:
    python fill_questionnaire.py <template.xlsx> <data.json> <output.xlsx>
"""
import sys, json, re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

# 実物の全角シート名対応
def ws_of(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    alt = name.replace('4', '４')
    return wb[alt] if alt in wb.sheetnames else None

def _cell(ws, coord):
    m = re.match(r'^([A-Z]+)(\d+)$', coord)
    return ws.cell(row=int(m.group(2)), column=column_index_from_string(m.group(1)))

def put(ws, coord, value, left=True):
    if value in (None, ''):
        return
    c = _cell(ws, coord)
    c.value = value
    c.alignment = Alignment(horizontal='left' if left else 'center',
                            vertical='center', wrap_text=True)

def fmt_date(iso):
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', str(iso or ''))
    return f"{int(m.group(1))}年 {int(m.group(2))}月 {int(m.group(3))}日" if m else (iso or '')

# ===== チェックボックス処理（入管書式: ■ で塗りつぶし。✓/☑ は使用禁止） =====
CHECKBOX_MARK = '■'
# テンプレートに現れ得る空チェックボックス系記号（先に来るものほど優先）
_CHECKBOX_EMPTY_TOKENS = ('□', '☐', '⬜︎', '⬜', '✓', '☑', '■')

def _replace_first_checkbox_token(text):
    """セル文字列中の最初のチェックボックス記号 1 つだけを '■' に置換。
    元のレイアウト（位置・ラベル文字）は維持する。"""
    s = str(text) if text is not None else ''
    for tok in _CHECKBOX_EMPTY_TOKENS:
        if tok in s:
            return s.replace(tok, CHECKBOX_MARK, 1)
    # チェックボックス記号が見当たらない場合は '■' を返す（空セル用フォールバック）
    return CHECKBOX_MARK

def check(ws, coord):
    """指定セルのチェックボックス記号（□/☐/⬜︎/⬜ 等）を '■' に置換。
    ※入管指定フォーマット: ✓ / ☑ は使用禁止。必ず ■ を使用する。
    ※枠の座標を変えないよう、セルの値だけを差し替える。"""
    c = _cell(ws, coord)
    if c.value:
        c.value = _replace_first_checkbox_token(c.value)
    else:
        # 元々空のセルに対しても ■ を入れる（座標は変えない）
        c.value = CHECKBOX_MARK

def circle(ws, coord, target):
    """セル内の特定文字を ○target に (役所書式の○囲み)"""
    c = _cell(ws, coord)
    if c.value and target in str(c.value):
        c.value = str(c.value).replace(target, '○' + target, 1)

def circle_word(ws, coord):
    """セル全体(単語のみ)を ○単語 に"""
    c = _cell(ws, coord)
    if c.value:
        c.value = '○' + str(c.value)

def textblock(ws, rng, text):
    """自由記述ブロック: 範囲結合して折返しテキスト"""
    if text in (None, ''):
        return
    min_c, min_r, max_c, max_r = range_boundaries(rng)
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= min_r and mr.max_row <= max_r and mr.min_col >= min_c and mr.max_col <= max_c:
            ws.unmerge_cells(str(mr))
    ws.merge_cells(rng)
    c = ws.cell(row=min_r, column=min_c)
    c.value = text
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

def name_stacked(ws, name_coord, furi_box, kana, kanji):
    """氏名: フリガナ(小・上)+漢字(下) の縦積み。隣のフリガナ枠ラベルは消す"""
    if not kanji and not kana:
        return
    c = _cell(ws, name_coord)
    if kana:
        c.value = CellRichText([TextBlock(InlineFont(rFont='ＭＳ 明朝', sz=8), kana + "\n"),
                                TextBlock(InlineFont(rFont='ＭＳ 明朝', sz=11), kanji)])
    else:
        c.value = kanji
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if furi_box:
        _cell(ws, furi_box).value = None

def g(d, key, default=''):
    return d.get(key, default)


# ============================================================
# Sheet1: 身分事項
# ============================================================
def fill_sheet1(ws, d):
    put(ws, 'H19', g(d, 'applicant_nationality'))
    name_stacked(ws, 'Q18', None, '', g(d, 'applicant_name'))
    sx = g(d, 'applicant_sex')
    if sx == '男': put(ws, 'AV18', '○男', left=False)
    elif sx == '女': put(ws, 'AV19', '○女', left=False)
    # 配偶者: D21の「氏名」は項目ラベルなので残す。Q21の塊にフリガナ(上)+氏名(下)
    name_stacked(ws, 'Q21', None, g(d, 'spouse_name_kana'), g(d, 'spouse_name_kanji'))
    put(ws, 'AN21', g(d, 'spouse_nationality'))
    put(ws, 'Q23', g(d, 'spouse_home_address'))
    if g(d, 'spouse_home_phone'): put(ws, 'Q24', f"自宅　{g(d,'spouse_home_phone')}")
    if g(d, 'spouse_mobile_phone'): put(ws, 'AE24', f"携帯　{g(d,'spouse_mobile_phone')}")
    co = g(d, 'cohabitant_exists', '無')
    if co == '無':
        put(ws, 'Q25', '■無　　□有　　（氏名　　　　　　　　　　　　　　　　　）')
    else:
        put(ws, 'Q25', f"□無　　■有　　（氏名 {g(d,'cohabitant_name')}）")
    ht = g(d, 'housing_type')
    if ht == '自己所有': put(ws, 'H26', '■', left=False); put(ws, 'O26', '□', left=False)
    elif ht == '借家': put(ws, 'H26', '□', left=False); put(ws, 'O26', '■', left=False)
    rent = g(d, 'housing_rent_amount'); ldk = g(d, 'housing_ldk')
    if rent or ldk:
        rent_s = f"{int(rent):,}" if isinstance(rent, (int, float)) else rent
        put(ws, 'U26', f"家賃 {rent_s} 円　{ldk}")
    put(ws, 'Q27', g(d, 'spouse_company_name'))
    put(ws, 'AM27', g(d, 'spouse_job_description'))
    put(ws, 'Q28', g(d, 'spouse_company_address'))
    put(ws, 'Q29', g(d, 'spouse_company_phone'))
    if g(d, 'spouse_employment_date'): put(ws, 'AM29', fmt_date(g(d, 'spouse_employment_date')))


# ============================================================
# Sheet2: 結婚に至った経緯
# ============================================================
def fill_sheet2(ws, d):
    if g(d, 'first_meeting_date'): put(ws, 'P6', fmt_date(g(d, 'first_meeting_date')))
    put(ws, 'P8', g(d, 'first_meeting_place'))
    textblock(ws, 'B17:AW36', g(d, 'relationship_history'))


# ============================================================
# Sheet3: 紹介者・夫婦間の言語
# ============================================================
def fill_sheet3(ws, d):
    if g(d, 'introducer_exists') == '有':
        check(ws, 'D5')
    else:
        check(ws, 'D4')
    put(ws, 'P7', g(d, 'introducer_nationality'))
    put(ws, 'P9', g(d, 'introducer_name'))
    sx = g(d, 'introducer_sex')
    if sx == '男': circle(ws, 'AD9', '男')
    elif sx == '女': circle(ws, 'AG9', '女')
    if g(d, 'introducer_birthdate'): put(ws, 'P11', fmt_date(g(d, 'introducer_birthdate')))
    put(ws, 'P13', g(d, 'introducer_address'))
    put(ws, 'P15', g(d, 'introducer_phone'))
    put(ws, 'Z17', g(d, 'introducer_residence_card_no'))
    if g(d, 'introduction_date'): put(ws, 'P21', fmt_date(g(d, 'introduction_date')))
    put(ws, 'P23', g(d, 'introduction_place'))
    meth = g(d, 'introduction_method')
    mmap = {'写真': 'P25', '電話': 'T25', '対面': 'X25',
            'E-mail': 'AB25', 'Ｅ－ｍａｉｌ': 'AB25', 'その他': 'P27'}
    if meth in mmap:
        check(ws, mmap[meth])
    elif meth:
        check(ws, 'P27'); put(ws, 'V27', meth)
    textblock(ws, 'B30:AI33', g(d, 'intro_relation_to_applicant'))
    textblock(ws, 'B35:AI38', g(d, 'intro_relation_to_spouse'))
    put(ws, 'D48', g(d, 'daily_conversation_language'))
    put(ws, 'R52', g(d, 'applicant_mother_tongue'))
    put(ws, 'R53', g(d, 'spouse_mother_tongue'))


# ============================================================
# Sheet4: 言語理解度・結婚届証人
# ============================================================
def _lang_cell(val, cells):
    if not val: return None
    if '難しい' in val or '通訳' in val: return cells[0]
    if '筆談' in val or 'あいさつ' in val: return cells[1]
    if '日常会話' in val: return cells[2]
    if '支障' in val: return cells[3]
    return None

def fill_sheet4(ws, d):
    au = _lang_cell(g(d, 'applicant_understands_spouse_lang'), ('E5', 'X5', 'E7', 'X7'))
    if au: check(ws, au)
    su = _lang_cell(g(d, 'spouse_understands_applicant_lang'), ('E12', 'X12', 'E14', 'X14'))
    if su: check(ws, su)
    textblock(ws, 'D19:AW23', g(d, 'applicant_japanese_learning'))
    textblock(ws, 'F28:AW32', g(d, 'communication_method'))
    put(ws, 'L36', g(d, 'interpreter_name'))
    put(ws, 'L38', g(d, 'interpreter_nationality'))
    put(ws, 'L40', g(d, 'interpreter_address'))
    put(ws, 'G45', g(d, 'witness1_name'))
    if g(d, 'witness1_sex') == '男': circle(ws, 'AQ45', '男')
    elif g(d, 'witness1_sex') == '女': circle(ws, 'AQ45', '女')
    put(ws, 'G47', g(d, 'witness1_address'))
    put(ws, 'G49', g(d, 'witness1_phone'))
    put(ws, 'G51', g(d, 'witness2_name'))
    if g(d, 'witness2_sex') == '男': circle(ws, 'AQ51', '男')
    elif g(d, 'witness2_sex') == '女': circle(ws, 'AQ51', '女')
    put(ws, 'G53', g(d, 'witness2_address'))
    put(ws, 'G55', g(d, 'witness2_phone'))


# ============================================================
# Sheet5: 結婚式・結婚歴・申請人の来日歴
# ============================================================
def fill_sheet5(ws, d):
    # 結婚式
    if g(d, 'wedding_date'): put(ws, 'G4', fmt_date(g(d, 'wedding_date')))
    put(ws, 'G6', g(d, 'wedding_place'))
    att = {'father': 'K', 'mother': 'O', 'brother': 'S', 'yng_bro': 'W',
           'sister': 'AA', 'yng_sis': 'AE', 'child': 'AI'}
    for k, col in att.items():
        if g(d, f'wedding_attendee_applicant_{k}'): circle(ws, f'{col}8', ws[f'{col}8'].value or '')
    # 出席者は単語のみセルなので circle_word を使う
    for k, col in att.items():
        if g(d, f'wedding_attendee_applicant_{k}'):
            c = _cell(ws, f'{col}8')
            if c.value and not str(c.value).startswith('○'): c.value = '○' + str(c.value)
        if g(d, f'wedding_attendee_spouse_{k}'):
            c = _cell(ws, f'{col}10')
            if c.value and not str(c.value).startswith('○'): c.value = '○' + str(c.value)
    if g(d, 'wedding_total_attendees'): put(ws, 'M12', str(g(d, 'wedding_total_attendees')))
    # 申請人 結婚歴 (□チェックボックス: 初婚=M17, 再婚=M19)
    if g(d, 'applicant_marriage_status') == '初婚':
        check(ws, 'M17')
    elif g(d, 'applicant_marriage_status') == '再婚':
        check(ws, 'M19')
        put(ws, 'Q19', str(g(d, 'applicant_remarriage_count')))
        if g(d, 'applicant_prev_marriage_start'): put(ws, 'M21', fmt_date(g(d, 'applicant_prev_marriage_start')))
        if g(d, 'applicant_prev_marriage_end'): put(ws, 'AC21', fmt_date(g(d, 'applicant_prev_marriage_end')))
        r = g(d, 'applicant_prev_marriage_reason')
        if r == '離婚': check(ws, 'AG23')   # （□離婚
        elif r == '死別': check(ws, 'AK23')  # □死別
    # 配偶者 結婚歴 (初婚=M25, 再婚=M27)
    if g(d, 'spouse_marriage_status') == '初婚':
        check(ws, 'M25')
    elif g(d, 'spouse_marriage_status') == '再婚':
        check(ws, 'M27')
        put(ws, 'Q27', str(g(d, 'spouse_remarriage_count')))
        if g(d, 'spouse_prev_marriage_start'): put(ws, 'M29', fmt_date(g(d, 'spouse_prev_marriage_start')))
        if g(d, 'spouse_prev_marriage_end'): put(ws, 'AC29', fmt_date(g(d, 'spouse_prev_marriage_end')))
        r = g(d, 'spouse_prev_marriage_reason')
        if r == '離婚': check(ws, 'AG31')
        elif r == '死別': check(ws, 'AK31')
    # 申請人 来日歴
    if g(d, 'applicant_japan_visits_count'): put(ws, 'F37', str(g(d, 'applicant_japan_visits_count')))
    rows = {1: 41, 2: 43, 3: 45, 4: 47, 5: 49}
    for i, row in rows.items():
        s = g(d, f'applicant_japan_visit_{i}_start')
        e = g(d, f'applicant_japan_visit_{i}_end')
        p = g(d, f'applicant_japan_visit_{i}_purpose')
        if s: put(ws, f'E{row}', fmt_date(s))
        if e: put(ws, f'T{row}', fmt_date(e))
        if p: put(ws, f'AI{row}', p)


# ============================================================
# Sheet6: 配偶者の渡航歴・退去強制歴
# ============================================================
def fill_sheet6(ws, d):
    # 結婚前の渡航
    if g(d, 'spouse_visit_before_marriage_count'): put(ws, 'Q2', str(g(d, 'spouse_visit_before_marriage_count')))
    bm = {1: 4, 2: 6, 3: 8, 4: 10, 5: 12}
    for i, row in bm.items():
        if g(d, f'spouse_visit_bm_{i}_start'): put(ws, f'E{row}', fmt_date(g(d, f'spouse_visit_bm_{i}_start')))
        if g(d, f'spouse_visit_bm_{i}_end'): put(ws, f'Z{row}', fmt_date(g(d, f'spouse_visit_bm_{i}_end')))
    # 結婚後の渡航
    if g(d, 'spouse_visit_after_marriage_count'): put(ws, 'I16', str(g(d, 'spouse_visit_after_marriage_count')))
    am = {1: 18, 2: 20, 3: 22, 4: 24, 5: 26}
    for i, row in am.items():
        if g(d, f'spouse_visit_am_{i}_start'): put(ws, f'E{row}', fmt_date(g(d, f'spouse_visit_am_{i}_start')))
        if g(d, f'spouse_visit_am_{i}_end'): put(ws, f'Z{row}', fmt_date(g(d, f'spouse_visit_am_{i}_end')))
    # 退去強制歴 (□チェックボックス: 無=C33, 有=C35)
    if g(d, 'deportation_exists') == '有':
        check(ws, 'C35')
        put(ws, 'F35', str(g(d, 'deportation_count')))
        if g(d, 'deportation_reason_overstay'): check(ws, 'D42')
        if g(d, 'deportation_reason_illegal_entry'): check(ws, 'D44')
        if g(d, 'deportation_reason_other'):
            check(ws, 'D46'); put(ws, 'I46', g(d, 'deportation_reason_other_text'))
        if g(d, 'deportation_date'): put(ws, 'I50', fmt_date(g(d, 'deportation_date')))
        put(ws, 'AC50', g(d, 'deportation_airport'))
        if g(d, 'deportation_passport_same') == '同じ':
            check(ws, 'D55')
        elif g(d, 'deportation_passport_same'):
            check(ws, 'D57')
            put(ws, 'N57', g(d, 'deportation_prev_nationality'))
            put(ws, 'N59', g(d, 'deportation_prev_name'))
            if g(d, 'deportation_prev_birthdate'): put(ws, 'N61', fmt_date(g(d, 'deportation_prev_birthdate')))
    else:
        check(ws, 'C33')


# ============================================================
# Sheet7: 親族情報
# ============================================================
def fill_sheet7(ws, d):
    # 退去強制までの同居(該当時)
    if g(d, 'deport_cohabit_period_start'): put(ws, 'I5', fmt_date(g(d, 'deport_cohabit_period_start')))
    if g(d, 'deport_cohabit_period_end'): put(ws, 'AC5', fmt_date(g(d, 'deport_cohabit_period_end')))
    put(ws, 'I7', g(d, 'deport_cohabit_address'))
    # 夫の親族 (父=18, 母=19, 兄弟姉妹=20-22)
    put(ws, 'G18', g(d, 'husband_father_name')); put(ws, 'U18', str(g(d, 'husband_father_age') or ''))
    put(ws, 'X18', g(d, 'husband_father_address')); put(ws, 'AO18', g(d, 'husband_father_phone'))
    put(ws, 'G19', g(d, 'husband_mother_name')); put(ws, 'U19', str(g(d, 'husband_mother_age') or ''))
    put(ws, 'X19', g(d, 'husband_mother_address')); put(ws, 'AO19', g(d, 'husband_mother_phone'))
    for i, row in {1: 20, 2: 21, 3: 22}.items():
        put(ws, f'D{row}', g(d, f'husband_sibling_{i}_relation'))
        put(ws, f'G{row}', g(d, f'husband_sibling_{i}_name'))
        put(ws, f'U{row}', str(g(d, f'husband_sibling_{i}_age') or ''))
        put(ws, f'X{row}', g(d, f'husband_sibling_{i}_address'))
        put(ws, f'AO{row}', g(d, f'husband_sibling_{i}_phone'))
    # 妻の親族 (父=27, 母=28, 兄弟姉妹=29-31)
    put(ws, 'G27', g(d, 'wife_father_name')); put(ws, 'U27', str(g(d, 'wife_father_age') or ''))
    put(ws, 'X27', g(d, 'wife_father_address')); put(ws, 'AO27', g(d, 'wife_father_phone'))
    put(ws, 'G28', g(d, 'wife_mother_name')); put(ws, 'U28', str(g(d, 'wife_mother_age') or ''))
    put(ws, 'X28', g(d, 'wife_mother_address')); put(ws, 'AO28', g(d, 'wife_mother_phone'))
    for i, row in {1: 29, 2: 30, 3: 31}.items():
        put(ws, f'D{row}', g(d, f'wife_sibling_{i}_relation'))
        put(ws, f'G{row}', g(d, f'wife_sibling_{i}_name'))
        put(ws, f'U{row}', str(g(d, f'wife_sibling_{i}_age') or ''))
        put(ws, f'X{row}', g(d, f'wife_sibling_{i}_address'))
        put(ws, f'AO{row}', g(d, f'wife_sibling_{i}_phone'))


# ============================================================
# Sheet8: お子さん・結婚の認知・署名
# ============================================================
def fill_sheet8(ws, d):
    # お子さん (記載例は行4,5。実入力は行6-10)
    for i, row in {1: 6, 2: 7, 3: 8, 4: 9, 5: 10}.items():
        put(ws, f'B{row}', g(d, f'child_{i}_relation'))
        put(ws, f'G{row}', g(d, f'child_{i}_name'))
        if g(d, f'child_{i}_birthdate'): put(ws, f'S{row}', fmt_date(g(d, f'child_{i}_birthdate')))
        put(ws, f'Z{row}', g(d, f'child_{i}_address'))
    # 結婚を知っている親族 (○で囲む)
    km = {'father': 'G', 'mother': 'J', 'brother': 'M', 'yng_bro': 'P',
          'sister': 'S', 'yng_sis': 'V', 'child': 'Y'}
    for k, col in km.items():
        if g(d, f'marriage_known_husband_{k}'):
            c = _cell(ws, f'{col}15')
            if c.value and not str(c.value).startswith('○'): c.value = '○' + str(c.value)
        if g(d, f'marriage_known_wife_{k}'):
            c = _cell(ws, f'{col}16')
            if c.value and not str(c.value).startswith('○'): c.value = '○' + str(c.value)
    # 署名
    if g(d, 'signature_year'): put(ws, 'F23', f"{g(d,'signature_year')}年", left=False)
    if g(d, 'signature_month'): put(ws, 'L23', f"{g(d,'signature_month')}月", left=False)
    if g(d, 'signature_day'): put(ws, 'R23', f"{g(d,'signature_day')}日", left=False)
    put(ws, 'N26', g(d, 'signature_spouse_name'))


# ============================================================
# メイン
# ============================================================

# --- 印刷設定（PDF/印刷レイアウトの固定）---
PAGE_SETUP = {
    'Sheet1': (None, True,  1,    1,    (0.5118110236, 0.5118110236, 0.3543307087, 0.7480314961, 0.3149606299, 0.3149606299)),
    'Sheet2': (76,   True,  None, None, (0.7,          0.7,          0.75,         0.75,         0.3,          0.3)),
    'Sheet3': (None, True,  1,    1,    (0.7086614173, 0.7086614173, 0.7480314961, 0.7480314961, 0.5118110236, 0.5118110236)),
    'Sheet4': (63,   True,  None, None, (0.7,          0.7,          0.75,         0.75,         0.5,          0.5)),
    'Sheet5': (None, True,  1,    1,    (0.7086614173, 0.7086614173, 0.5511811024, 0.1574803150, 0.5118110236, 0.5118110236)),
    'Sheet6': (64,   True,  None, None, (0.1181102362, 0.0,          0.3937007874, 0.0,          0.5118110236, 0.5118110236)),
    'Sheet7': (93,   True,  None, None, (0.1181102362, 0.1181102362, 0.3543307087, 0.1574803150, 0.5118110236, 0.5118110236)),
    # 🌟 Sheet8: 左右余白1.3cm (0.511811インチ)、すべての行を1ページに印刷 (fit_w=0, fit_h=1)
    'Sheet8': (None, True,  0,    1,    (0.5118110236, 0.5118110236, 0.75,         0.75,         0.5,          0.5)),
}


def apply_page_setup(wb):
    """印刷・PDF出力時のレイアウトを固定する。"""
    for name, (scale, fit_page, fit_w, fit_h, margins) in PAGE_SETUP.items():
        ws = ws_of(wb, name)
        if ws is None:
            continue
        ps = ws.page_setup
        ps.orientation = 'portrait'
        ps.paperSize = 9  # 🌟 A4サイズに強制固定 (Adobe PDFエラー対策)
        ps.scale = scale
        ps.fitToWidth = fit_w
        ps.fitToHeight = fit_h
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=bool(fit_page))
        
        l, r, t, b, h, f = margins
        ws.page_margins.left, ws.page_margins.right = l, r
        ws.page_margins.top, ws.page_margins.bottom = t, b
        ws.page_margins.header, ws.page_margins.footer = h, f
        
        # 印刷範囲を解除
        ws.print_area = None

    # 🌟 Adobe PDFエラー対策: ブック全体に残った壊れた印刷範囲を完全に削除
    try:
        for key in list(wb.defined_names.keys()):
            if 'Print_Area' in key or 'Print_Titles' in key:
                del wb.defined_names[key]
    except Exception:
        pass


def format_fixes(wb):
    """実Excel表示用の列幅調整など (フォームの見切れ対策)"""
    def widen(ws, cols, width):
        from copy import copy
        from openpyxl.utils import column_index_from_string
        for c in cols:
            if c in ws.column_dimensions:
                # 共有オブジェクトによるXML重複出力を防ぐため、コピーして単一列に分離
                cd = copy(ws.column_dimensions[c])
                idx = column_index_from_string(c)
                cd.min = idx
                cd.max = idx
                cd.width = width
                ws.column_dimensions[c] = cd
            else:
                ws.column_dimensions[c].width = width
    
    s4 = ws_of(wb, 'Sheet4')
    if s4 is not None: widen(s4, ['C', 'D', 'E'], 4.2)
    s5 = ws_of(wb, 'Sheet5')
    if s5 is not None: widen(s5, ['C', 'D', 'E'], 4.2)
    s7 = ws_of(wb, 'Sheet7')
    if s7 is not None:
        widen(s7, ['B', 'C'], 3.6)
        widen(s7, ['D', 'E', 'F'], 4.2)
        # 死亡表示用に住所列(X,Y)を少しだけ広げる(「死」だけ表示される問題対策)
        widen(s7, ['X', 'Y'], 3.0)
        # クライアント要望: 16,17,25,26行目の続柄欄(B〜F列)の縦幅を伸ばす
        for r in (16, 17, 25, 26):
            current = s7.row_dimensions[r].height or 15
            s7.row_dimensions[r].height = max(current, 22)

    # シート3: ベース列幅2.2、性別列(AD〜AH)は幅3
    s3 = ws_of(wb, 'Sheet3')
    if s3 is not None:
        import string
        cols_s3 = list(string.ascii_uppercase) + ['A' + c for c in string.ascii_uppercase[:15]]
        widen(s3, cols_s3, 2.2)
        widen(s3, ['AD', 'AE', 'AF', 'AG', 'AH'], 3)

def _delete_cell(ws, coord):
    """セルを内部ストレージから完全削除し、使用範囲(dimension)を縮める"""
    from openpyxl.utils.cell import coordinate_to_tuple
    ws._cells.pop(coordinate_to_tuple(coord), None)


def cleanup_strays(wb):
    """A4枠の外に残った空白セル・結合を除去する（2ページ化の原因をなくす）。"""
    from openpyxl.utils import column_index_from_string
    s1 = ws_of(wb, 'Sheet1')
    if s1 is not None:
        for coord in ('BE25', 'BF26'):
            _delete_cell(s1, coord)
    s5 = ws_of(wb, 'Sheet5')
    if s5 is not None:
        ba = column_index_from_string('BA')
        for mr in list(s5.merged_cells.ranges):
            if mr.max_col >= ba:
                s5.unmerge_cells(str(mr))
        for coord in ('BA20', 'BA29'):
            _delete_cell(s5, coord)
    s3 = ws_of(wb, 'Sheet3')
    if s3 is not None:
        for rng in ('AD9:AE9', 'AG9:AH9'):
            try:
                s3.unmerge_cells(rng)
            except Exception:
                pass
        for coord in list(s3._cells.keys()):
            if coord[1] > 41:
                s3._cells.pop(coord, None)
        for col in list(s3.column_dimensions.keys()):
            if col.isalpha() and column_index_from_string(col) > 41:
                del s3.column_dimensions[col]


def trim_trailing_empty(ws):
    """中身(値・罫線・塗り)の右端より右にある空セル・列幅設定を削除し、
    使用範囲を実際の内容に合わせる（余分な空セルによる横方向の2ページ化を防ぐ）。
    """
    from openpyxl.utils import column_index_from_string
    content_max = 0
    for row in ws.iter_rows():
        for c in row:
            has_val = c.value is not None and str(c.value).strip() != ''
            has_bd = c.border and any([c.border.left.style, c.border.right.style,
                                       c.border.top.style, c.border.bottom.style])
            has_fl = c.fill and c.fill.patternType is not None
            if (has_val or has_bd or has_fl) and c.column > content_max:
                content_max = c.column
    if content_max == 0:
        return
    for key in [k for k in list(ws._cells.keys()) if k[1] > content_max]:
        ws._cells.pop(key, None)
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col > content_max:
            ws.unmerge_cells(str(mr))
    for col in list(ws.column_dimensions.keys()):
        if col.isalpha() and column_index_from_string(col) > content_max:
            del ws.column_dimensions[col]


# ==========================================================
# 👇 Sheet8の完璧なフォーマット調整（列幅をExcel上で1.5にする） 👇
# ==========================================================
def apply_sheet8_perfect_formatting(ws):
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string

    LAST = 37  # AK列（=フォームの実際の右端。住所セルは Z〜AK の結合）。これより右は消す

    # 1. AL列より右の「結合・セル・列幅」をすべて物理削除する
    for mr in list(ws.merged_cells.ranges):
        if mr.max_col > LAST:
            ws.unmerge_cells(str(mr))
    for key in [k for k in list(ws._cells.keys()) if k[1] > LAST]:
        ws._cells.pop(key, None)
    for col_key in list(ws.column_dimensions.keys()):
        cd = ws.column_dimensions[col_key]
        try:
            idx = column_index_from_string(col_key)
        except Exception:
            idx = getattr(cd, 'min', 0) or 0
        cmax = getattr(cd, 'max', idx) or idx
        if idx > LAST or cmax > LAST:
            del ws.column_dimensions[col_key]

    # 2. A〜AK の列幅をExcel上「1.5」(=Python側 2.25) に。1列ずつ分離設定(XML重複防止)
    for col_idx in range(1, LAST + 1):
        L = get_column_letter(col_idx)
        cd = ws.column_dimensions[L]
        cd.min = col_idx
        cd.max = col_idx
        cd.width = 2.25

    # 3. 表の結合と罫線の再構築 (3行目〜10行目)
    thin = Side(style='thin', color='000000')
    double = Side(style='double', color='000000')

    col_ranges = [
        (2, 6),   # B〜F: 続柄
        (7, 18),  # G〜R: 氏名
        (19, 25), # S〜Y: 生年月日
        (26, 37)  # Z〜AK: 住所
    ]

    for row in range(3, 11):
        # 既存の結合を解除 (B〜AKの範囲)
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row == row and mr.max_row == row and mr.min_col >= 2 and mr.max_col <= 37:
                ws.unmerge_cells(str(mr))
        
        # 再結合
        for min_col, max_col in col_ranges:
            ws.merge_cells(start_row=row, start_column=min_col, end_row=row, end_column=max_col)

        # 罫線の設定
        if row == 3 or row == 5:
            bottom_side = double
        else:
            bottom_side = thin
            
        if row == 3:
            top_side = thin
        elif row == 4 or row == 6:
            top_side = double
        else:
            top_side = thin

        for min_col, max_col in col_ranges:
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                left_side = thin if col == min_col else Side(style=None)
                right_side = thin if col == max_col else Side(style=None)
                
                cell.border = Border(
                    left=left_side,
                    right=right_side,
                    top=top_side,
                    bottom=bottom_side
                )

    # 👇👇👇 ここから差し込む 👇👇👇
    # 34行目 (A34:AK34) の結合と中央揃え
    from openpyxl.styles import Alignment
    
    # 既存の「- 8/8 -」などのテキストを探して保持する
    footer_text = '- 8/8 -'
    for c in range(1, 38):
        val = ws.cell(row=34, column=c).value
        if val and '8/8' in str(val):
            footer_text = val
            break

    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == 34 and mr.max_row == 34:
            ws.unmerge_cells(str(mr)) # 念のため既存の34行目の結合を解除
    ws.merge_cells('A34:AK34')
    ws['A34'].value = footer_text  # ★保持したテキストを再セット
    ws['A34'].alignment = Alignment(horizontal='center', vertical='center')
    # 👆👆👆 ここまで差し込む 👆👆👆

    # 4. 印刷範囲をA1〜AK（フォームの右端）に固定。
    ws.print_area = f'A1:AK{ws.max_row}'


def _fill_workbook(wb, data):

    """workbookにデータを転記して整形 (fill_all / fill_to_bytes 共通処理)"""
    fillers = {
        'Sheet1': fill_sheet1, 'Sheet2': fill_sheet2, 'Sheet3': fill_sheet3,
        'Sheet4': fill_sheet4, 'Sheet5': fill_sheet5, 'Sheet6': fill_sheet6,
        'Sheet7': fill_sheet7, 'Sheet8': fill_sheet8,
    }
    fd = data.get('form_data', data)
    for name, fn in fillers.items():
        ws = ws_of(wb, name)
        if ws is not None:
            fn(ws, fd)

    cleanup_strays(wb)     # 枠外の空白セル・結合を除去

    for _ws in wb.worksheets:
        trim_trailing_empty(_ws)   # 各シートの余分な空セルを削除

    apply_page_setup(wb)   # 印刷倍率の固定・1ページ化
    format_fixes(wb)       # 全体微調整

    # 🌟 Sheet8の最強フォーマット（列幅1.5 ＆ 右枠線引き直し）を一番最後に適用！
    # （trim_trailing_empty等で消されるのを防ぐため、すべての処理の最後に実行します）
    ws8 = ws_of(wb, 'Sheet8')
    if ws8:
        apply_sheet8_perfect_formatting(ws8)

    # 全シートの目盛線(gridlines)を非表示に（提出様式は枠線で表現するため）
    for _ws in wb.worksheets:
        _ws.sheet_view.showGridLines = False

    return wb


def fill_all(template_path, data, output_path):
    wb = load_workbook(template_path)
    _fill_workbook(wb, data)
    wb.save(output_path)
    return output_path


def normalize_borders(wb):
    """隣接セルで二重に引かれた同じ罫線を1本に整える。
    （上セルの下線と下セルの上線が両方あると、Excelの画面表示で太線に見えるため）
    結合セルの内側の線はExcelが描画しない仕様のため、
    「実際に描画される線どうしが重なっている場合」だけを1本にする。"""
    from openpyxl.styles import Border, Side
    for ws in wb.worksheets:
        # 結合範囲マップ: そのセルの各辺が実際に描画されるか（範囲の外周だけ描画される）
        edge = {}
        for rng in ws.merged_cells.ranges:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    edge[(r, c)] = (c == rng.min_col, c == rng.max_col,
                                    r == rng.min_row, r == rng.max_row)  # L,R,T,B

        def renders(r, c, side):
            e = edge.get((r, c))
            if e is None:
                return True
            L, R, T, B = e
            return {"left": L, "right": R, "top": T, "bottom": B}[side]

        min_r, min_c = ws.min_row, ws.min_column
        rows = list(ws.iter_rows())
        for r_i, row in enumerate(rows):
            for c_i, cell in enumerate(row):
                b = cell.border
                r, c = min_r + r_i, min_c + c_i
                new_top, new_left = b.top, b.left
                changed = False
                if b.top and b.top.style and r_i > 0:
                    ab = rows[r_i - 1][c_i].border
                    if (ab.bottom and ab.bottom.style == b.top.style
                            and renders(r, c, "top") and renders(r - 1, c, "bottom")):
                        new_top = Side(style=None); changed = True
                if b.left and b.left.style and c_i > 0:
                    lb = rows[r_i][c_i - 1].border
                    if (lb.right and lb.right.style == b.left.style
                            and renders(r, c, "left") and renders(r, c - 1, "right")):
                        new_left = Side(style=None); changed = True
                if changed:
                    cell.border = Border(
                        left=new_left, right=b.right, top=new_top, bottom=b.bottom,
                        diagonal=b.diagonal, diagonal_direction=b.diagonal_direction,
                        diagonalUp=b.diagonalUp, diagonalDown=b.diagonalDown,
                    )


def fill_to_bytes(template_path, data):
    """Streamlit等で使用: 転記済みExcelをbytesで返す"""
    import io
    wb = load_workbook(template_path)
    _fill_workbook(wb, data)
    normalize_borders(wb)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


if __name__ == '__main__':
    template, datafile, output = sys.argv[1], sys.argv[2], sys.argv[3]
    with open(datafile, encoding='utf-8') as f:
        data = json.load(f)
    fill_all(template, data, output)
    print(f'✅ 転記完了: {output}')