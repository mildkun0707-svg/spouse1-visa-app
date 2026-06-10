"""
🌸 配偶者ビザ申請 ヒアリングシートアプリ
================================================
- 行政書士がサイドバーで「ケースフラグ」をON/OFF
- クライアントが必要な項目だけ入力
- 入力データはJSONとして保存 → openpyxlでExcelに自動転記
================================================
"""
import streamlit as st
import json
from pathlib import Path
from datetime import date, datetime
from collections import defaultdict

# ============================================
# 設定
# ============================================
st.set_page_config(
    page_title="配偶者ビザ ヒアリングシート",
    page_icon="💞",
    layout="wide",
)

# ============================================
# 多言語化（翻訳）設定
# ============================================
TRANSLATIONS = {
    "en": {
        "入力済み項目": "Completed items",
        # --- 基本情報 ---
        "申請人(お相手)の国籍・地域": "Nationality / Region of Applicant",
        "申請人(お相手)の氏名(パスポート表記)": "Name of Applicant (as in Passport)",
        "申請人の性別": "Gender of Applicant",
        "申請人の生年月日": "Date of Birth of Applicant",
        "配偶者(あなた)の氏名": "Name of Spouse (You)",
        "配偶者(あなた)のフリガナ": "Furigana of Spouse (You)",
        "配偶者(あなた)の国籍・地域": "Nationality / Region of Spouse",
        "配偶者の自宅住所": "Home Address of Spouse",
        "配偶者の自宅電話番号": "Home Phone Number of Spouse",
        "配偶者の携帯電話番号": "Mobile Phone Number of Spouse",
        "同居者の有無": "Co-residents",
        "同居者の氏名(同居者がいる場合)": "Name of Co-residents (if any)",
        "住居形態": "Housing Type",
        "家賃(円)": "Rent (Yen)",
        "間取り(LDK)": "Floor Plan (e.g., LDK)",
        "配偶者の勤務先 会社名": "Spouse's Workplace (Company Name)",
        "配偶者の職務内容": "Spouse's Job Description",
        "配偶者の勤務先 所在地": "Spouse's Workplace Address",
        "配偶者の勤務先 電話番号": "Spouse's Workplace Phone Number",
        "配偶者の就職年月日": "Spouse's Date of Employment",
        # --- 出会い・紹介 ---
        "初めて会った時期(年月日)": "When you first met (Date)",
        "初めて会った場所": "Where you first met",
        "出会いから結婚届提出までの経緯(できるだけ詳しく)": "History from first meeting to marriage registration (in detail)",
        "紹介者の有無": "Presence of an Introducer",
        "紹介者の国籍": "Nationality of Introducer",
        "紹介者の氏名(または会社名)": "Name of Introducer (or Company)",
        "紹介者の性別": "Gender of Introducer",
        "紹介者の生年月日": "Date of Birth of Introducer",
        "紹介者の住所": "Address of Introducer",
        "紹介者の電話番号": "Phone Number of Introducer",
        "紹介者が外国人の場合 在留カード番号": "Residence Card Number (if Introducer is foreign)",
        "紹介された年月日": "Date of Introduction",
        "紹介された場所": "Place of Introduction",
        "紹介方法": "Method of Introduction",
        "「その他」の場合の具体内容": "Details if 'Other'",
        "紹介者と申請人の関係(詳しく)": "Relationship between Introducer and Applicant (in detail)",
        "紹介者と配偶者の関係(詳しく)": "Relationship between Introducer and Spouse (in detail)",
        # --- 言語・通訳 ---
        "日常の夫婦の会話に使われている言語": "Language used in daily conversation",
        "申請人の母(国)語": "Applicant's Mother Tongue",
        "配偶者の母(国)語": "Spouse's Mother Tongue",
        "申請人は配偶者の母国語をどの程度理解できるか": "How well does the applicant understand the spouse's mother tongue?",
        "配偶者は申請人の母国語をどの程度理解できるか": "How well does the spouse understand the applicant's mother tongue?",
        "申請人が日本語を学んだ経緯(具体的に)": "How the applicant learned Japanese (specifically)",
        "言葉が通じない場合の意思疎通方法": "Method of communication if languages are not understood",
        "通訳者の氏名": "Name of Interpreter",
        "通訳者の国籍": "Nationality of Interpreter",
        "通訳者の住所": "Address of Interpreter",
        # --- 証人・結婚式 ---
        "証人①氏名": "Witness 1: Name", "証人①性別": "Witness 1: Gender", "証人①住所": "Witness 1: Address", "証人①電話番号": "Witness 1: Phone Number",
        "証人②氏名": "Witness 2: Name", "証人②性別": "Witness 2: Gender", "証人②住所": "Witness 2: Address", "証人②電話番号": "Witness 2: Phone Number",
        "結婚式・披露宴の年月日": "Date of Wedding Ceremony / Reception",
        "結婚式の場所": "Location of Wedding Ceremony",
        "申請人側出席者:父": "Applicant's Attendees: Father", "申請人側出席者:母": "Applicant's Attendees: Mother",
        "申請人側出席者:兄": "Applicant's Attendees: Older Brother", "申請人側出席者:弟": "Applicant's Attendees: Younger Brother",
        "申請人側出席者:姉": "Applicant's Attendees: Older Sister", "申請人側出席者:妹": "Applicant's Attendees: Younger Sister", "申請人側出席者:子": "Applicant's Attendees: Child",
        "配偶者側出席者:父": "Spouse's Attendees: Father", "配偶者側出席者:母": "Spouse's Attendees: Mother",
        "配偶者側出席者:兄": "Spouse's Attendees: Older Brother", "配偶者側出席者:弟": "Spouse's Attendees: Younger Brother",
        "配偶者側出席者:姉": "Spouse's Attendees: Older Sister", "配偶者側出席者:妹": "Spouse's Attendees: Younger Sister", "配偶者側出席者:子": "Spouse's Attendees: Child",
        "双方の出席者 合計(人)": "Total Attendees (Persons)",
        # --- 結婚歴 ---
        "申請人の結婚歴": "Applicant's Marriage History",
        "申請人が再婚の場合:何回目": "If Applicant is remarrying: Which time?",
        "申請人の前回結婚 開始日": "Applicant's Previous Marriage: Start Date",
        "申請人の前回結婚 終了日": "Applicant's Previous Marriage: End Date",
        "申請人の前婚解消理由": "Reason for dissolution of Applicant's previous marriage",
        "配偶者の結婚歴": "Spouse's Marriage History",
        "配偶者が再婚の場合:何回目": "If Spouse is remarrying: Which time?",
        "配偶者の前回結婚 開始日": "Spouse's Previous Marriage: Start Date",
        "配偶者の前回結婚 終了日": "Spouse's Previous Marriage: End Date",
        "配偶者の前婚解消理由": "Reason for dissolution of Spouse's previous marriage",
        # --- 渡航歴 ---
        "申請人のこれまでの来日回数": "Number of times Applicant has visited Japan",
        "来日歴1 開始日": "Visit to Japan 1: Start Date", "来日歴1 終了日": "Visit to Japan 1: End Date", "来日歴1 目的": "Visit to Japan 1: Purpose",
        "来日歴2 開始日": "Visit to Japan 2: Start Date", "来日歴2 終了日": "Visit to Japan 2: End Date", "来日歴2 目的": "Visit to Japan 2: Purpose",
        "来日歴3 開始日": "Visit to Japan 3: Start Date", "来日歴3 終了日": "Visit to Japan 3: End Date", "来日歴3 目的": "Visit to Japan 3: Purpose",
        "来日歴4 開始日": "Visit to Japan 4: Start Date", "来日歴4 終了日": "Visit to Japan 4: End Date", "来日歴4 目的": "Visit to Japan 4: Purpose",
        "来日歴5 開始日": "Visit to Japan 5: Start Date", "来日歴5 終了日": "Visit to Japan 5: End Date", "来日歴5 目的": "Visit to Japan 5: Purpose",
        "結婚前の渡航回数": "Number of visits before marriage",
        "結婚前渡航1 開始日": "Pre-marriage Visit 1: Start Date", "結婚前渡航1 終了日": "Pre-marriage Visit 1: End Date",
        "結婚前渡航2 開始日": "Pre-marriage Visit 2: Start Date", "結婚前渡航2 終了日": "Pre-marriage Visit 2: End Date",
        "結婚前渡航3 開始日": "Pre-marriage Visit 3: Start Date", "結婚前渡航3 終了日": "Pre-marriage Visit 3: End Date",
        "結婚前渡航4 開始日": "Pre-marriage Visit 4: Start Date", "結婚前渡航4 終了日": "Pre-marriage Visit 4: End Date",
        "結婚前渡航5 開始日": "Pre-marriage Visit 5: Start Date", "結婚前渡航5 終了日": "Pre-marriage Visit 5: End Date",
        "結婚後の渡航回数": "Number of visits after marriage",
        "結婚後渡航1 開始日": "Post-marriage Visit 1: Start Date", "結婚後渡航1 終了日": "Post-marriage Visit 1: End Date",
        "結婚後渡航2 開始日": "Post-marriage Visit 2: Start Date", "結婚後渡航2 終了日": "Post-marriage Visit 2: End Date",
        "結婚後渡航3 開始日": "Post-marriage Visit 3: Start Date", "結婚後渡航3 終了日": "Post-marriage Visit 3: End Date",
        "結婚後渡航4 開始日": "Post-marriage Visit 4: Start Date", "結婚後渡航4 終了日": "Post-marriage Visit 4: End Date",
        "結婚後渡航5 開始日": "Post-marriage Visit 5: Start Date", "結婚後渡航5 終了日": "Post-marriage Visit 5: End Date",
        # --- 退去強制歴 ---
        "申請人の退去強制歴の有無": "History of Deportation of Applicant",
        "退去強制された回数": "Number of times deported",
        "違反内容:不法残留(オーバーステイ)": "Violation: Illegal Overstay",
        "違反内容:不法入国": "Violation: Illegal Entry",
        "違反内容:その他": "Violation: Other",
        "その他の具体内容": "Details of Other Violation",
        "退去強制された年月日(直近)": "Date of Deportation (Most recent)",
        "出国した空港名": "Airport of Departure",
        "当時のパスポート情報と今回の申請情報": "Passport info at that time vs Current application info",
        "当時の国籍": "Nationality at that time",
        "当時の氏名": "Name at that time",
        "当時の生年月日": "Date of Birth at that time",
        "退去強制までに同居した期間(開始)": "Period of cohabitation before deportation (Start)",
        "退去強制までに同居した期間(終了)": "Period of cohabitation before deportation (End)",
        "同居の住所": "Address of cohabitation",
        # --- 親族情報 ---
        "夫の父:氏名": "Husband's Father: Name", "夫の父:年齢": "Husband's Father: Age", "夫の父:住所": "Husband's Father: Address", "夫の父:電話番号": "Husband's Father: Phone",
        "夫の母:氏名": "Husband's Mother: Name", "夫の母:年齢": "Husband's Mother: Age", "夫の母:住所": "Husband's Mother: Address", "夫の母:電話番号": "Husband's Mother: Phone",
        "夫の兄弟姉妹1:続柄": "Husband's Sibling 1: Relation", "夫の兄弟姉妹1:氏名": "Husband's Sibling 1: Name", "夫の兄弟姉妹1:年齢": "Husband's Sibling 1: Age", "夫の兄弟姉妹1:住所": "Husband's Sibling 1: Address", "夫の兄弟姉妹1:電話番号": "Husband's Sibling 1: Phone",
        "夫の兄弟姉妹2:続柄": "Husband's Sibling 2: Relation", "夫の兄弟姉妹2:氏名": "Husband's Sibling 2: Name", "夫の兄弟姉妹2:年齢": "Husband's Sibling 2: Age", "夫の兄弟姉妹2:住所": "Husband's Sibling 2: Address", "夫の兄弟姉妹2:電話番号": "Husband's Sibling 2: Phone",
        "夫の兄弟姉妹3:続柄": "Husband's Sibling 3: Relation", "夫の兄弟姉妹3:氏名": "Husband's Sibling 3: Name", "夫の兄弟姉妹3:年齢": "Husband's Sibling 3: Age", "夫の兄弟姉妹3:住所": "Husband's Sibling 3: Address", "夫の兄弟姉妹3:電話番号": "Husband's Sibling 3: Phone",
        "妻の父:氏名": "Wife's Father: Name", "妻の父:年齢": "Wife's Father: Age", "妻の父:住所": "Wife's Father: Address", "妻の父:電話番号": "Wife's Father: Phone",
        "妻の母:氏名": "Wife's Mother: Name", "妻の母:年齢": "Wife's Mother: Age", "妻の母:住所": "Wife's Mother: Address", "妻の母:電話番号": "Wife's Mother: Phone",
        "妻の兄弟姉妹1:続柄": "Wife's Sibling 1: Relation", "妻の兄弟姉妹1:氏名": "Wife's Sibling 1: Name", "妻の兄弟姉妹1:年齢": "Wife's Sibling 1: Age", "妻の兄弟姉妹1:住所": "Wife's Sibling 1: Address", "妻の兄弟姉妹1:電話番号": "Wife's Sibling 1: Phone",
        "妻の兄弟姉妹2:続柄": "Wife's Sibling 2: Relation", "妻の兄弟姉妹2:氏名": "Wife's Sibling 2: Name", "妻の兄弟姉妹2:年齢": "Wife's Sibling 2: Age", "妻の兄弟姉妹2:住所": "Wife's Sibling 2: Address", "妻の兄弟姉妹2:電話番号": "Wife's Sibling 2: Phone",
        "妻の兄弟姉妹3:続柄": "Wife's Sibling 3: Relation", "妻の兄弟姉妹3:氏名": "Wife's Sibling 3: Name", "妻の兄弟姉妹3:年齢": "Wife's Sibling 3: Age", "妻の兄弟姉妹3:住所": "Wife's Sibling 3: Address", "妻の兄弟姉妹3:電話番号": "Wife's Sibling 3: Phone",
        "お子さん1:続柄": "Child 1: Relation", "お子さん1:氏名": "Child 1: Name", "お子さん1:生年月日": "Child 1: Date of Birth", "お子さん1:住所": "Child 1: Address",
        "お子さん2:続柄": "Child 2: Relation", "お子さん2:氏名": "Child 2: Name", "お子さん2:生年月日": "Child 2: Date of Birth", "お子さん2:住所": "Child 2: Address",
        "お子さん3:続柄": "Child 3: Relation", "お子さん3:氏名": "Child 3: Name", "お子さん3:生年月日": "Child 3: Date of Birth", "お子さん3:住所": "Child 3: Address",
        "お子さん4:続柄": "Child 4: Relation", "お子さん4:氏名": "Child 4: Name", "お子さん4:生年月日": "Child 4: Date of Birth", "お子さん4:住所": "Child 4: Address",
        "お子さん5:続柄": "Child 5: Relation", "お子さん5:氏名": "Child 5: Name", "お子さん5:生年月日": "Child 5: Date of Birth", "お子さん5:住所": "Child 5: Address",
        "夫側で結婚を知っている:父": "Husband's side knows about marriage: Father", "夫側で結婚を知っている:母": "Husband's side knows about marriage: Mother",
        "夫側で結婚を知っている:兄": "Husband's side knows about marriage: Older Brother", "夫側で結婚を知っている:弟": "Husband's side knows about marriage: Younger Brother",
        "夫側で結婚を知っている:姉": "Husband's side knows about marriage: Older Sister", "夫側で結婚を知っている:妹": "Husband's side knows about marriage: Younger Sister", "夫側で結婚を知っている:子": "Husband's side knows about marriage: Child",
        "妻側で結婚を知っている:父": "Wife's side knows about marriage: Father", "妻側で結婚を知っている:母": "Wife's side knows about marriage: Mother",
        "妻側で結婚を知っている:兄": "Wife's side knows about marriage: Older Brother", "妻側で結婚を知っている:弟": "Wife's side knows about marriage: Younger Brother",
        "妻側で結婚を知っている:姉": "Wife's side knows about marriage: Older Sister", "妻側で結婚を知っている:妹": "Wife's side knows about marriage: Younger Sister", "妻側で結婚を知っている:子": "Wife's side knows about marriage: Child",
        "署名年": "Signature Year", "署名月": "Signature Month", "署名日": "Signature Day", "配偶者の署名": "Signature of Spouse",
        # --- 選択肢・プレースホルダー・ヘルプテキスト ---
        "男": "Male", "女": "Female", "無": "None", "有": "Yes", "自己所有": "Owned", "借家": "Rented",
        "写真": "Photo", "電話": "Phone", "対面": "In-person", "E-mail": "E-mail", "その他": "Other",
        "難しい(通訳必要)": "Difficult (Interpreter needed)", "筆談/あいさつ程度": "Written / Greetings only", "日常会話程度は可能": "Daily conversation possible", "会話に支障なし": "No problem in conversation",
        "初婚": "First Marriage", "再婚": "Remarriage", "離婚": "Divorced", "死別": "Bereaved", "同じ": "Same", "別の氏名等": "Different Name, etc.",
        "例: ベトナム、フィリピン、中国": "e.g., Vietnam, Philippines, China", "例: NGUYEN THI HOA": "e.g., NGUYEN THI HOA",
        "例: 山田 太郎": "e.g., Taro Yamada", "例: ヤマダ タロウ": "e.g., Yamada Taro", "例: 日本": "e.g., Japan",
        "例: 東京都新宿区西新宿2-8-1": "e.g., 2-8-1 Nishi-Shinjuku, Shinjuku-ku, Tokyo", "例: 03-1234-5678": "e.g., 03-1234-5678", "例: 090-1234-5678": "e.g., 090-1234-5678",
        "例: 85000": "e.g., 85000", "例: 2LDK": "e.g., 2LDK", "例: 株式会社サンプル商事": "e.g., Sample Trading Co., Ltd.", "例: 営業職": "e.g., Sales",
        "例: 東京都渋谷区のカフェ": "e.g., Cafe in Shibuya, Tokyo", "例: 日本語、英語、中国語": "e.g., Japanese, English, Chinese", "例: ベトナム語": "e.g., Vietnamese",
        "例: 観光、仕事、留学": "e.g., Sightseeing, Work, Study", "例: 兄、姉、弟、妹": "e.g., Older brother, Older sister, Younger brother, Younger sister", "例: 夫の長男、妻の長女": "e.g., Husband's eldest son, Wife's eldest daughter",
        "同居者ありを選んだ場合のみ入力": "Enter only if 'Yes' is selected for Co-residents", "借家を選んだ場合のみ": "Only if 'Rented' is selected",
        "年月日を示しながら詳しく記載": "Describe in detail, including dates", "結婚相談所による紹介の場合は会社名": "Company name if introduced by a marriage agency", "死亡の場合は「死亡」と記載": "Write 'Deceased' if applicable",
        # --- グループ名・サイドバー ---
        "①身分事項(基本情報)": "1. Basic Information",
        "②結婚に至った経緯": "2. History of Marriage",
        "③-A 紹介者": "3-A. Introducer",
        "③-B 夫婦間の言語": "3-B. Language between Spouses",
        "④-A 言語理解度": "4-A. Language Comprehension",
        "④-B 通訳": "4-B. Interpreter",
        "④-C 結婚届証人(日本国内婚のみ)": "4-C. Marriage Witnesses (Only in Japan)",
        "⑤-A 結婚式・披露宴": "5-A. Wedding Ceremony / Reception",
        "⑤-B 結婚歴": "5-B. Marriage History",
        "⑤-C 申請人の来日歴": "5-C. Applicant's Visits to Japan",
        "⑥-A 配偶者の渡航歴(結婚前)": "6-A. Spouse's Overseas Visits (Before Marriage)",
        "⑥-B 配偶者の渡航歴(結婚後)": "6-B. Spouse's Overseas Visits (After Marriage)",
        "⑥-C 退去強制歴": "6-C. History of Deportation",
        "⑦-A 退去強制後の同居": "7-A. Cohabitation after Deportation",
        "⑦-B 夫の親族": "7-B. Husband's Relatives",
        "⑦-C 妻の親族": "7-C. Wife's Relatives",
        "⑧-A お子さん": "8-A. Children",
        "⑧-B 結婚を知っている親族": "8-B. Relatives who know about the marriage",
        "⑧-C 署名": "8-C. Signature",
        "紹介者がいる(お見合い結婚・結婚相談所等)": "Has an Introducer (Arranged marriage, Agency, etc.)",
        "結婚式・披露宴を行った": "Held a Wedding Ceremony / Reception",
        "申請人(お相手)が再婚": "Applicant is remarrying",
        "配偶者(あなたのクライアント)が再婚": "Spouse (Client) is remarrying",
        "申請人の来日歴あり": "Applicant has visited Japan before",
        "配偶者の申請人母国訪問歴あり": "Spouse has visited Applicant's home country",
        "退去強制歴あり": "Has a history of deportation",
        "お子さんあり": "Has children",
        "日本国内で婚姻届出した": "Registered marriage in Japan",
        "通訳が必要(言語が通じない)": "Needs an interpreter (Language barrier)",
        "申請人が日本語を学習した経験あり": "Applicant has studied Japanese",
        "紹介者が外国人": "Introducer is a foreign national",
        # --- アプリUI（ヘッダー・ボタン・注意書き）---
        "配偶者ビザ申請 ヒアリングシート": "Spouse Visa Application — Hearing Sheet",
        "全{n}項目": "All {n} items",
        "⚠️ 入力内容は自動保存されません": "⚠️ Your entries are NOT saved automatically",
        "作業を中断する場合は、下部の「💾 入力データをJSON保存」ボタンで必ずデータを保存してください。次回そのJSONを読み込めば続きから入力できます。":
            "If you need to pause, be sure to save your data using the “💾 Save input data as JSON” button at the bottom. You can load that JSON next time to continue where you left off.",
        "質問書の作成": "Create the Questionnaire",
        "✨ 質問書Excelを作成する": "✨ Create Questionnaire (Excel)",
        "📥 質問書Excelをダウンロード": "📥 Download Questionnaire (Excel)",
        "💾 入力データをJSON保存": "💾 Save input data as JSON",
        "管理者ログイン": "Admin login",
        "パスワード": "Password",
        "ログイン": "Log in",
        "パスワードが違います": "Incorrect password",
        "先生に送信する": "Send to your immigration lawyer",
        "送信しています…": "Sending…",
        "✅ 先生に送信しました。ご入力ありがとうございました。": "✅ Sent to your lawyer. Thank you for your input!",
        "⬇️ 送信できなかった場合：データを保存して先生にメールで送ってください": "⬇️ If sending failed: save the data and email it to your lawyer",
        "（氏名未入力）": "(name not entered)",
        "入力が終わったら、いちばん下の「📨 先生に送信する」ボタンを押してください。入力内容がそのまま先生に届きます。": "When you finish, press the \"📨 Send to your immigration lawyer\" button at the bottom. Your answers will be delivered directly to your lawyer.",
        "📊 入力済み項目を確認": "📊 Review entered items",
        "🔄 入力をリセット": "🔄 Reset all inputs",
        "入力済みの内容": "Entered content",
        "✅ 質問書を作成しました。下のボタンからダウンロードできます。": "✅ Questionnaire created. Download it with the button below.",
        "🗑️ 入力内容をすべてリセットしました。": "🗑️ All inputs have been reset.",
        "⚙️ ケース設定(行政書士用)": "⚙️ Case Settings (for the administrative scrivener)",
         "該当する項目にチェックを入れると、関連する質問が表示されます": "Check the items that apply; the related questions will appear.",
        "表示中の項目数": "Items currently shown",
        "表示項目数": "Items shown"
    },
    "zh": {
        "入力済み項目": "已填写的项目",
        # --- 基本情報 ---
        "申請人(お相手)の国籍・地域": "申请人（对方）的国籍/地区",
        "申請人(お相手)の氏名(パスポート表記)": "申请人姓名（与护照一致）",
        "申請人の性別": "申请人的性别",
        "申請人の生年月日": "申请人出生日期",
        "配偶者(あなた)の氏名": "配偶者（您）的姓名",
        "配偶者(あなた)のフリガナ": "配偶者（您）的片假名注音",
        "配偶者(あなた)の国籍・地域": "配偶者的国籍/地区",
        "配偶者の自宅住所": "配偶者的家庭住址",
        "配偶者の自宅電話番号": "配偶者的家庭电话",
        "配偶者の携帯電話番号": "配偶者的手机号码",
        "同居者の有無": "是否有同住人",
        "同居者の氏名(同居者がいる場合)": "同住人姓名（如有）",
        "住居形態": "居住形态",
        "家賃(円)": "房租（日元）",
        "間取り(LDK)": "户型（如LDK）",
        "配偶者の勤務先 会社名": "配偶者的工作单位（公司名称）",
        "配偶者的職務内容": "配偶者的工作内容",
        "配偶者の勤務先 所在地": "配偶者的工作单位地址",
        "配偶者の勤務先 電話番号": "配偶者的工作单位电话",
        "配偶者の就職年月日": "配偶者的入职日期",
        # --- 出会い・紹介 ---
        "初めて会った時期(年月日)": "初次见面的时间（年月日）",
        "初めて会った場所": "初次见面的地点",
        "出会いから結婚届提出までの経緯(できるだけ詳しく)": "从相识到提交结婚登记的经过（尽可能详细）",
        "紹介者の有無": "是否有介绍人",
        "紹介者の国籍": "介绍人的国籍",
        "紹介者の氏名(または会社名)": "介绍人的姓名（或公司名称）",
        "紹介者の性別": "介绍人的性别",
        "紹介者の生年月日": "介绍人的出生日期",
        "紹介者の住所": "介绍人的地址",
        "紹介者の電話番号": "介绍人的电话号码",
        "紹介者が外国人の場合 在留カード番号": "介绍人为外国人时的在留卡号",
        "紹介された年月日": "被介绍的日期",
        "紹介された場所": "被介绍的地点",
        "紹介方法": "介绍方式",
        "「その他」の場合の具体内容": "选择“其他”时的具体内容",
        "紹介者と申請人の関係(詳しく)": "介绍人与申请人的关系（详细）",
        "紹介者と配偶者の関係(詳しく)": "介绍人与配偶者的关系（详细）",
        # --- 言語・通訳 ---
        "日常の夫婦の会話に使われている言語": "夫妻间日常交流使用的语言",
        "申請人の母(国)語": "申请人的母语",
        "配偶者の母(国)語": "配偶者的母语",
        "申請人は配偶者の母国語をどの程度理解できるか": "申请人对配偶者母语的理解程度",
        "配偶者は申請人の母国語をどの程度理解できるか": "配偶者对申请人母语的理解程度",
        "申請人が日本語を学んだ経緯(具体的に)": "申请人学习日语的经过（具体）",
        "言葉が通じない場合の意思疎通方法": "语言不通时的沟通方式",
        "通訳者の氏名": "翻译人员姓名",
        "通訳者の国籍": "翻译人员国籍",
        "通訳者の住所": "翻译人员地址",
        # --- 証人・結婚式 ---
        "証人①氏名": "证人①姓名", "証人①性別": "证人①性别", "証人①住所": "证人①地址", "証人①電話番号": "证人①电话号码",
        "証人②氏名": "证人②姓名", "証人②性別": "证人②性别", "証人②住所": "证人②地址", "証人②電話番号": "证人②电话号码",
        "結婚式・披露宴の年月日": "婚礼/婚宴的日期",
        "結婚式の場所": "婚礼地点",
        "申請人側出席者:父": "申请人方出席者：父亲", "申請人側出席者:母": "申请人方出席者：母亲",
        "申請人側出席者:兄": "申请人方出席者：哥哥", "申請人側出席者:弟": "申请人方出席者：弟弟",
        "申請人側出席者:姉": "申请人方出席者：姐姐", "申請人側出席者:妹": "申请人方出席者：妹妹", "申請人側出席者:子": "申请人方出席者：子女",
        "配偶者側出席者:父": "配偶者方出席者：父亲", "配偶者側出席者:母": "配偶者方出席者：母亲",
        "配偶者側出席者:兄": "配偶者方出席者：哥哥", "配偶者側出席者:弟": "配偶者方出席者：弟弟",
        "配偶者側出席者:姉": "配偶者方出席者：姐姐", "配偶者側出席者:妹": "配偶者方出席者：妹妹", "配偶者側出席者:子": "配偶者方出席者：子女",
        "双方の出席者 合計(人)": "双方出席者总计（人）",
        # --- 結婚歴 ---
        "申請人の結婚歴": "申请人的婚姻史",
        "申請人が再婚の場合:何回目": "申请人若为再婚：第几次？",
        "申請人の前回結婚 開始日": "申请人上一段婚姻：开始日期",
        "申請人の前回結婚 終了日": "申请人上一段婚姻：结束日期",
        "申請人の前婚解消理由": "申请人上一段婚姻解除理由",
        "配偶者の結婚歴": "配偶者的婚姻史",
        "配偶者が再婚の場合:何回目": "配偶者若为再婚：第几次？",
        "配偶者の前回結婚 開始日": "配偶者上一段婚姻：开始日期",
        "配偶者の前回結婚 終了日": "配偶者上一段婚姻：结束日期",
        "配偶者の前婚解消理由": "配偶者上一段婚姻解除理由",
        # --- 渡航歴 ---
        "申請人のこれまでの来日回数": "申请人迄今为止的来日次数",
        "来日歴1 開始日": "来日记录1：开始日期", "来日歴1 終了日": "来日记录1：结束日期", "来日歴1 目的": "来日记录1：目的",
        "来日歴2 開始日": "来日记录2：开始日期", "来日歴2 終了日": "来日记录2：结束日期", "来日歴2 目的": "来日记录2：目的",
        "来日歴3 開始日": "来日记录3：开始日期", "来日歴3 終了日": "来日记录3：结束日期", "来日歴3 目的": "来日记录3：目的",
        "来日歴4 開始日": "来日记录4：开始日期", "来日歴4 終了日": "来日记录4：结束日期", "来日歴4 目的": "来日记录4：目的",
        "来日歴5 開始日": "来日记录5：开始日期", "来日歴5 終了日": "来日记录5：结束日期", "来日歴5 目的": "来日记录5：目的",
        "結婚前の渡航回数": "婚前出国次数",
        "結婚前渡航1 開始日": "婚前出国1：开始日期", "結婚前渡航1 終了日": "婚前出国1：结束日期",
        "結婚前渡航2 開始日": "婚前出国2：开始日期", "結婚前渡航2 終了日": "婚前出国2：结束日期",
        "結婚前渡航3 開始日": "婚前出国3：开始日期", "結婚前渡航3 終了日": "婚前出国3：结束日期",
        "結婚前渡航4 開始日": "婚前出国4：开始日期", "結婚前渡航4 終了日": "婚前出国4：结束日期",
        "結婚前渡航5 開始日": "婚前出国5：开始日期", "結婚前渡航5 終了日": "婚前出国5：结束日期",
        "結婚後の渡航回数": "婚后出国次数",
        "結婚後渡航1 開始日": "婚后出国1：开始日期", "結婚後渡航1 終了日": "婚后出国1：结束日期",
        "結婚後渡航2 開始日": "婚后出国2：开始日期", "結婚後渡航2 終了日": "婚后出国2：结束日期",
        "結婚後渡航3 開始日": "婚后出国3：开始日期", "結婚後渡航3 終了日": "婚后出国3：结束日期",
        "結婚後渡航4 開始日": "婚后出国4：开始日期", "結婚後渡航4 終了日": "婚后出国4：结束日期",
        "結婚後渡航5 開始日": "婚后出国5：开始日期", "結婚後渡航5 終了日": "婚后出国5：结束日期",
        # --- 退去強制歴 ---
        "申請人の退去強制歴の有無": "申请人是否有被遣返记录",
        "退去強制された回数": "被遣返次数",
        "違反内容:不法残留(オーバーステイ)": "违规内容：非法滞留（逾期居留）",
        "違反内容:不法入国": "违规内容：非法入境",
        "違反内容:その他": "违规内容：其他",
        "その他の具体内容": "其他的具体内容",
        "退去強制された年月日(直近)": "被遣返的日期（最近一次）",
        "出国した空港名": "离境机场名称",
        "当時のパスポート情報と今回の申請情報": "当时的护照信息与本次申请信息",
        "当時の国籍": "当时的国籍",
        "当時の氏名": "当时的姓名",
        "当時の生年月日": "当时的出生日期",
        "退去強制までに同居した期間(開始)": "被遣返前同居的期间（开始）",
        "退去強制までに同居した期間(終了)": "被遣返前同居的期间（结束）",
        "同居の住所": "同居地址",
        # --- 親族情報 ---
        "夫の父:氏名": "丈夫的父亲：姓名", "夫の父:年齢": "丈夫的父亲：年龄", "夫の父:住所": "丈夫的父亲：地址", "夫の父:電話番号": "丈夫的父亲：电话号码",
        "夫の母:氏名": "丈夫的母亲：姓名", "夫の母:年齢": "丈夫的母亲：年龄", "夫の母:住所": "丈夫的母亲：地址", "夫の母:電話番号": "丈夫的母亲：电话号码",
        "夫の兄弟姉妹1:続柄": "丈夫的兄弟姐妹1：关系", "夫の兄弟姉妹1:氏名": "丈夫的兄弟姐妹1：姓名", "夫の兄弟姉妹1:年齢": "丈夫的兄弟姐妹1：年龄", "夫の兄弟姉妹1:住所": "丈夫的兄弟姐妹1：地址", "夫の兄弟姉妹1:電話番号": "丈夫的兄弟姐妹1：电话号码",
        "夫の兄弟姉妹2:続柄": "丈夫的兄弟姐妹2：关系", "夫の兄弟姉妹2:氏名": "丈夫的兄弟姐妹2：姓名", "夫の兄弟姉妹2:年齢": "丈夫的兄弟姐妹2：年龄", "夫の兄弟姉妹2:住所": "丈夫的兄弟姐妹2：地址", "夫の兄弟姉妹2:電話番号": "丈夫的兄弟姐妹2：电话号码",
        "夫の兄弟姉妹3:続柄": "丈夫的兄弟姐妹3：关系", "夫の兄弟姉妹3:氏名": "丈夫的兄弟姐妹3：姓名", "夫の兄弟姉妹3:年齢": "丈夫的兄弟姐妹3：年龄", "夫の兄弟姉妹3:住所": "丈夫的兄弟姐妹3：地址", "夫の兄弟姉妹3:電話番号": "丈夫的兄弟姐妹3：电话号码",
        "妻の父:氏名": "妻子的父亲：姓名", "妻の父:年齢": "妻子的父亲：年龄", "妻の父:住所": "妻子的父亲：地址", "妻の父:電話番号": "妻子的父亲：电话号码",
        "妻の母:氏名": "妻子的母亲：姓名", "妻の母:年齢": "妻子的母亲：年龄", "妻の母:住所": "妻子的母亲：地址", "妻の母:電話番号": "妻子的母亲：电话号码",
        "妻の兄弟姉妹1:続柄": "妻子的兄弟姐妹1：关系", "妻の兄弟姉妹1:氏名": "妻子的兄弟姐妹1：姓名", "妻の兄弟姉妹1:年齢": "妻子的兄弟姐妹1：年龄", "妻の兄弟姉妹1:住所": "妻子的兄弟姐妹1：地址", "妻の兄弟姉妹1:電話番号": "妻子的兄弟姐妹1：电话号码",
        "妻の兄弟姉妹2:続柄": "妻子的兄弟姐妹2：关系", "妻の兄弟姉妹2:氏名": "妻子的兄弟姐妹2：姓名", "妻の兄弟姉妹2:年齢": "妻子的兄弟姐妹2：年龄", "妻の兄弟姉妹2:住所": "妻子的兄弟姐妹2：地址", "妻の兄弟姉妹2:電話番号": "妻子的兄弟姐妹2：电话号码",
        "妻の兄弟姉妹3:続柄": "妻子的兄弟姐妹3：关系", "妻の兄弟姉妹3:氏名": "妻子的兄弟姐妹3：姓名", "妻の兄弟姉妹3:年齢": "妻子的兄弟姐妹3：年龄", "妻の兄弟姉妹3:住所": "妻子的兄弟姐妹3：地址", "妻の兄弟姉妹3:電話番号": "妻子的兄弟姐妹3：电话号码",
        "お子さん1:続柄": "子女1：关系", "お子さん1:氏名": "子女1：姓名", "お子さん1:生年月日": "子女1：出生日期", "お子さん1:住所": "子女1：地址",
        "お子さん2:続柄": "子女2：关系", "お子さん2:氏名": "子女2：姓名", "お子さん2:生年月日": "子女2：出生日期", "お子さん2:住所": "子女2：地址",
        "お子さん3:続柄": "子女3：关系", "お子さん3:氏名": "子女3：姓名", "お子さん3:生年月日": "子女3：出生日期", "お子さん3:住所": "子女3：地址",
        "お子さん4:続柄": "子女4：关系", "お子さん4:氏名": "子女4：姓名", "お子さん4:生年月日": "子女4：出生日期", "お子さん4:住所": "子女4：地址",
        "お子さん5:続柄": "子女5：关系", "お子さん5:氏名": "子女5：姓名", "お子さん5:生年月日": "子女5：出生日期", "お子さん5:住所": "子女5：地址",
        "夫側で結婚を知っている:父": "男方知晓结婚的亲属：父亲", "夫側で結婚を知っている:母": "男方知晓结婚的亲属：母亲",
        "夫側で結婚を知っている:兄": "男方知晓结婚的亲属：哥哥", "夫側で結婚を知っている:弟": "男方知晓结婚的亲属：弟弟",
        "夫側で結婚を知っている:姉": "男方知晓结婚的亲属：姐姐", "夫側で結婚を知っている:妹": "男方知晓结婚的亲属：妹妹", "夫側で結婚を知っている:子": "男方知晓结婚的亲属：子女",
        "妻側で結婚を知っている:父": "女方知晓结婚的亲属：父亲", "妻側で結婚を知っている:母": "女方知晓结婚的亲属：母亲",
        "妻側で結婚を知っている:兄": "女方知晓结婚的亲属：哥哥", "妻側で結婚を知っている:弟": "女方知晓结婚的亲属：弟弟",
        "妻側で結婚を知っている:姉": "女方知晓结婚的亲属：姐姐", "妻側で結婚を知っている:妹": "女方知晓结婚的亲属：妹妹", "妻側で結婚を知っている:子": "女方知晓结婚的亲属：子女",
        "署名年": "签名年份", "署名月": "签名月份", "署名日": "签名日期", "配偶者の署名": "配偶者签名",
        # --- 選択肢・プレースホルダー・ヘルプテキスト ---
        "男": "男", "女": "女", "無": "无", "有": "有", "自己所有": "自有", "借家": "租借",
        "写真": "照片", "電話": "电话", "対面": "面谈", "E-mail": "电子邮件", "その他": "其他",
        "難しい(通訳必要)": "困难（需要翻译）", "筆談/あいさつ程度": "仅限笔谈/打招呼", "日常会話程度は可能": "可进行日常对话", "会話に支障なし": "交流无障碍",
        "初婚": "初婚", "再婚": "再婚", "離婚": "离婚", "死別": "丧偶", "同じ": "相同", "別の氏名等": "其他姓名等",
        "例: ベトナム、フィリピン、中国": "例：越南、菲律宾、中国", "例: NGUYEN THI HOA": "例：NGUYEN THI HOA",
        "例: 山田 太郎": "例：山田 太郎", "例: ヤマダ タロウ": "例：YAMADA TARO", "例: 日本": "例：日本",
        "例: 東京都新宿区西新宿2-8-1": "例：东京都新宿区西新宿2-8-1", "例: 03-1234-5678": "例：03-1234-5678", "例: 090-1234-5678": "例：090-1234-5678",
        "例: 85000": "例：85000", "例: 2LDK": "例：2LDK", "例: 株式会社サンプル商事": "例：Sample商事株式会社", "例: 営業職": "例：销售",
        "例: 東京都渋谷区のカフェ": "例：东京都涩谷区的咖啡馆", "例: 日本語、英語、中国語": "例：日语、英语、中文", "例: ベトナム語": "例：越南语",
        "例: 観光、仕事、留学": "例：旅游、工作、留学", "例: 兄、姉、弟、妹": "例：哥哥、姐姐、弟弟、妹妹", "例: 夫の長男、妻の長女": "例：丈夫的长子、妻子的长女",
        "同居者ありを選んだ場合のみ入力": "仅在选择“有同住人”时填写", "借家を選んだ場合のみ": "仅在选择“租借”时填写",
        "年月日を示しながら詳しく記載": "请注明年月日并详细填写", "結婚相談所による紹介の場合は会社名": "若是婚介所介绍，请填写公司名称", "死亡の場合は「死亡」と記載": "若已故，请填写“死亡”",
        # --- グループ名・サイドバー ---
        "①身分事項(基本情報)": "1. 身份事项（基本信息）",
        "②結婚に至った経緯": "2. 结婚经过",
        "③-A 紹介者": "3-A. 介绍人",
        "③-B 夫婦間の言語": "3-B. 夫妻间的语言",
        "④-A 言語理解度": "4-A. 语言理解程度",
        "④-B 通訳": "4-B. 翻译",
        "④-C 結婚届証人(日本国内婚のみ)": "4-C. 结婚登记证人（仅限日本国内结婚）",
        "⑤-A 結婚式・披露宴": "5-A. 婚礼・婚宴",
        "⑤-B 結婚歴": "5-B. 婚姻史",
        "⑤-C 申請人の来日歴": "5-C. 申请人来日记录",
        "⑥-A 配偶者の渡航歴(結婚前)": "6-A. 配偶者出国记录（婚前）",
        "⑥-B 配偶者の渡航歴(結婚後)": "6-B. 配偶者出国记录（婚后）",
        "⑥-C 退去強制歴": "6-C. 遣返记录",
        "⑦-A 退去強制後の同居": "7-A. 遣返后的同居情况",
        "⑦-B 夫の親族": "7-B. 丈夫的亲属",
        "⑦-C 妻の親族": "7-C. 妻子的亲属",
        "⑧-A お子さん": "8-A. 子女",
        "⑧-B 結婚を知っている親族": "8-B. 知晓结婚的亲属",
        "⑧-C 署名": "8-C. 签名",
        "紹介者がいる(お見合い結婚・結婚相談所等)": "有介绍人（相亲、婚介所等）",
        "結婚式・披露宴を行った": "举办了婚礼/婚宴",
        "申請人(お相手)が再婚": "申请人（对方）为再婚",
        "配偶者(あなたのクライアント)が再婚": "配偶者（您的客户）为再婚",
        "申請人の来日歴あり": "申请人有来日记录",
        "配偶者の申請人母国訪問歴あり": "配偶者曾访问过申请人母国",
        "退去強制歴あり": "有被遣返记录",
        "お子さんあり": "有子女",
        "日本国内で婚姻届出した": "在日本国内提交了结婚登记",
        "通訳が必要(言語が通じない)": "需要翻译（语言不通）",
        "申請人が日本語を学習した経験あり": "申请人有学习日语的经历",
        "紹介者が外国人": "介绍人为外国人",
        # --- アプリUI（ヘッダー・ボタン・注意書き）---
        "配偶者ビザ申請 ヒアリングシート": "配偶签证申请 问询表",
        "全{n}項目": "共{n}个项目",
        "⚠️ 入力内容は自動保存されません": "⚠️ 输入内容不会自动保存",
        "作業を中断する場合は、下部の「💾 入力データをJSON保存」ボタンで必ずデータを保存してください。次回そのJSONを読み込めば続きから入力できます。":
            "如需中途暂停，请务必使用底部的“💾 将输入数据保存为JSON”按钮保存数据。下次载入该JSON即可继续填写。",
        "質問書の作成": "制作质问书",
        "✨ 質問書Excelを作成する": "✨ 生成质问书（Excel）",
        "📥 質問書Excelをダウンロード": "📥 下载质问书（Excel）",
        "💾 入力データをJSON保存": "💾 将输入数据保存为JSON",
        "管理者ログイン": "管理员登录",
        "パスワード": "密码",
        "ログイン": "登录",
        "パスワードが違います": "密码错误",
        "先生に送信する": "发送给行政书士",
        "送信しています…": "正在发送…",
        "✅ 先生に送信しました。ご入力ありがとうございました。": "✅ 已发送给行政书士。感谢您的填写！",
        "⬇️ 送信できなかった場合：データを保存して先生にメールで送ってください": "⬇️ 如果发送失败：请下载数据并通过邮件发送给行政书士",
        "（氏名未入力）": "（未填写姓名）",
        "入力が終わったら、いちばん下の「📨 先生に送信する」ボタンを押してください。入力内容がそのまま先生に届きます。": "填写完成后，请点击最下方的「📨 发送给行政书士」按钮。填写内容将直接发送给行政书士。",
        "📊 入力済み項目を確認": "📊 查看已填写项目",
        "🔄 入力をリセット": "🔄 重置所有输入",
        "入力済みの内容": "已填写的内容",
        "✅ 質問書を作成しました。下のボタンからダウンロードできます。": "✅ 质问书已生成。可通过下方按钮下载。",
        "🗑️ 入力内容をすべてリセットしました。": "🗑️ 已重置全部输入内容。",
        "⚙️ ケース設定(行政書士用)": "⚙️ 案件设置（行政书士用）",
        "該当する項目にチェックを入れると、相关问题即会显示。": "勾选适用的项目，相关问题即会显示。",
        "表示中の項目数": "当前显示的项目数",
        "表示項目数": "显示项目数"
    },
    "vi": {
        "入力済み項目": "Mục đã hoàn thành",
        # --- 基本情報 ---
        "申請人(お相手)の国籍・地域": "Quốc tịch / Khu vực của Người nộp đơn",
        "申請人(お相手)の氏名(パスポート表記)": "Tên người nộp đơn (như trong Hộ chiếu)",
        "申請人の性別": "Giới tính của Người nộp đơn",
        "申請人の生年月日": "Ngày sinh của người nộp đơn",
        "配偶者(あなた)の氏名": "Tên của Vợ/Chồng (Bạn)",
        "配偶者(あなた)のフリガナ": "Furigana của Vợ/Chồng (Bạn)",
        "配偶者(あなた)の国籍・地域": "Quốc tịch / Khu vực của Vợ/Chồng",
        "配偶者の自宅住所": "Địa chỉ nhà của Vợ/Chồng",
        "配偶者の自宅電話番号": "Số điện thoại nhà của Vợ/Chồng",
        "配偶者の携帯電話番号": "Số điện thoại di động của Vợ/Chồng",
        "同居者の有無": "Người sống cùng",
        "同居者の氏名(同居者がいる場合)": "Tên người sống cùng (nếu có)",
        "住居形態": "Loại hình nhà ở",
        "家賃(円)": "Tiền thuê nhà (Yên)",
        "間取り(LDK)": "Sơ đồ mặt bằng (VD: LDK)",
        "配偶者の勤務先 会社名": "Nơi làm việc của Vợ/Chồng (Tên công ty)",
        "配偶者の職務内容": "Mô tả công việc của Vợ/Chồng",
        "配偶者の勤務先 所在地": "Địa chỉ nơi làm việc của Vợ/Chồng",
        "配偶者の勤務先 電話番号": "Số điện thoại nơi làm việc của Vợ/Chồng",
        "配偶者の就職年月日": "Ngày bắt đầu làm việc của Vợ/Chồng",
        # --- 出会い・紹介 ---
        "初めて会った時期(年月日)": "Thời gian gặp nhau lần đầu (Ngày tháng năm)",
        "初めて会った場所": "Địa điểm gặp nhau lần đầu",
        "出会いから結婚届提出までの経緯(できるだけ詳しく)": "Quá trình từ khi gặp nhau đến khi đăng ký kết hôn (chi tiết)",
        "紹介者の有無": "Có người giới thiệu không",
        "紹介者の国籍": "Quốc tịch của người giới thiệu",
        "紹介者の氏名(または会社名)": "Tên người giới thiệu (hoặc Tên công ty)",
        "紹介者の性別": "Giới tính của người giới thiệu",
        "紹介者の生年月日": "Ngày sinh của người giới thiệu",
        "紹介者の住所": "Địa chỉ của người giới thiệu",
        "紹介者の電話番号": "Số điện thoại của người giới thiệu",
        "紹介者が外国人の場合 在留カード番号": "Số thẻ lưu trú (nếu người giới thiệu là người nước ngoài)",
        "紹介された年月日": "Ngày được giới thiệu",
        "紹介された場所": "Địa điểm được giới thiệu",
        "紹介方法": "Phương thức giới thiệu",
        "「その他」の場合の具体内容": "Chi tiết nếu chọn 'Khác'",
        "紹介者と申請人の関係(詳しく)": "Mối quan hệ giữa người giới thiệu và người nộp đơn (chi tiết)",
        "紹介者と配偶者の関係(詳しく)": "Mối quan hệ giữa người giới thiệu và Vợ/Chồng (chi tiết)",
        # --- 言語・通訳 ---
        "日常の夫婦の会話に使われている言語": "Ngôn ngữ giao tiếp hàng ngày của vợ chồng",
        "申請人の母(国)語": "Ngôn ngữ mẹ đẻ của người nộp đơn",
        "配偶者の母(国)語": "Ngôn ngữ mẹ đẻ của Vợ/Chồng",
        "申請人は配偶者の母国語をどの程度理解できるか": "Người nộp đơn hiểu ngôn ngữ mẹ đẻ của Vợ/Chồng ở mức độ nào?",
        "配偶者は申請人の母国語をどの程度理解できるか": "Vợ/Chồng hiểu ngôn ngữ mẹ đẻ của người nộp đơn ở mức độ nào?",
        "申請人が日本語を学んだ経緯(具体的に)": "Quá trình người nộp đơn học tiếng Nhật (cụ thể)",
        "言葉が通じない場合の意思疎通方法": "Phương pháp giao tiếp khi không hiểu ngôn ngữ",
        "通訳者の氏名": "Tên người phiên dịch",
        "通訳者の国籍": "Quốc tịch của người phiên dịch",
        "通訳者の住所": "Địa chỉ của người phiên dịch",
        # --- 証人・結婚式 ---
        "証人①氏名": "Người làm chứng 1: Họ tên", "証人①性別": "Người làm chứng 1: Giới tính", "証人①住所": "Người làm chứng 1: Địa chỉ", "証人①電話番号": "Người làm chứng 1: Số điện thoại",
        "証人②氏名": "Người làm chứng 2: Họ tên", "証人②性別": "Người làm chứng 2: Giới tính", "証人②住所": "Người làm chứng 2: Địa chỉ", "証人②電話番号": "Người làm chứng 2: Số điện thoại",
        "結婚式・披露宴の年月日": "Ngày tổ chức lễ cưới / tiệc cưới",
        "結婚式の場所": "Địa điểm tổ chức lễ cưới",
        "申請人側出席者:父": "Khách mời phía người nộp đơn: Bố", "申請人側出席者:母": "Khách mời phía người nộp đơn: Mẹ",
        "申請人側出席者:兄": "Khách mời phía người nộp đơn: Anh trai", "申請人側出席者:弟": "Khách mời phía người nộp đơn: Em trai",
        "申請人側出席者:姉": "Khách mời phía người nộp đơn: Chị gái", "申請人側出席者:妹": "Khách mời phía người nộp đơn: Em gái", "申請人側出席者:子": "Khách mời phía người nộp đơn: Con",
        "配偶者側出席者:父": "Khách mời phía Vợ/Chồng: Bố", "配偶者側出席者:母": "Khách mời phía Vợ/Chồng: Mẹ",
        "配偶者側出席者:兄": "Khách mời phía Vợ/Chồng: Anh trai", "配偶者側出席者:弟": "Khách mời phía Vợ/Chồng: Em trai",
        "配偶者側出席者:姉": "Khách mời phía Vợ/Chồng: Chị gái", "配偶者側出席者:妹": "Khách mời phía Vợ/Chồng: Em gái", "配偶者側出席者:子": "Khách mời phía Vợ/Chồng: Con",
        "双方の出席者 合計(人)": "Tổng số khách mời (Người)",
        # --- 結婚歴 ---
        "申請人の結婚歴": "Lịch sử hôn nhân của người nộp đơn",
        "申請人が再婚の場合:何回目": "Nếu người nộp đơn tái hôn: Lần thứ mấy?",
        "申請人の前回結婚 開始日": "Hôn nhân trước của người nộp đơn: Ngày bắt đầu",
        "申請人の前回結婚 終了日": "Hôn nhân trước của người nộp đơn: Ngày kết thúc",
        "申請人の前婚解消理由": "Lý do chấm dứt hôn nhân trước của người nộp đơn",
        "配偶者の結婚歴": "Lịch sử hôn nhân của Vợ/Chồng",
        "配偶者が再婚の場合:何回目": "Nếu Vợ/Chồng tái hôn: Lần thứ mấy?",
        "配偶者の前回結婚 開始日": "Hôn nhân trước của Vợ/Chồng: Ngày bắt đầu",
        "配偶者の前回結婚 終了日": "Hôn nhân trước của Vợ/Chồng: Ngày kết thúc",
        "配偶者の前婚解消理由": "Lý do chấm dứt hôn nhân trước của Vợ/Chồng",
        # --- 渡航歴 ---
        "申請人のこれまでの来日回数": "Số lần người nộp đơn đã đến Nhật Bản",
        "来日歴1 開始日": "Lịch sử đến Nhật 1: Ngày bắt đầu", "来日歴1 終了日": "Lịch sử đến Nhật 1: Ngày kết thúc", "来日歴1 目的": "Lịch sử đến Nhật 1: Mục đích",
        "来日歴2 開始日": "Lịch sử đến Nhật 2: Ngày bắt đầu", "来日歴2 終了日": "Lịch sử đến Nhật 2: Ngày kết thúc", "来日歴2 目的": "Lịch sử đến Nhật 2: Mục đích",
        "来日歴3 開始日": "Lịch sử đến Nhật 3: Ngày bắt đầu", "来日歴3 終了日": "Lịch sử đến Nhật 3: Ngày kết thúc", "来日歴3 目的": "Lịch sử đến Nhật 3: Mục đích",
        "来日歴4 開始日": "Lịch sử đến Nhật 4: Ngày bắt đầu", "来日歴4 終了日": "Lịch sử đến Nhật 4: Ngày kết thúc", "来日歴4 目的": "Lịch sử đến Nhật 4: Mục đích",
        "来日歴5 開始日": "Lịch sử đến Nhật 5: Ngày bắt đầu", "来日歴5 終了日": "Lịch sử đến Nhật 5: Ngày kết thúc", "来日歴5 目的": "Lịch sử đến Nhật 5: Mục đích",
        "結婚前の渡航回数": "Số lần xuất ngoại trước khi kết hôn",
        "結婚前渡航1 開始日": "Xuất ngoại trước kết hôn 1: Ngày bắt đầu", "結婚前渡航1 終了日": "Xuất ngoại trước kết hôn 1: Ngày kết thúc",
        "結婚前渡航2 開始日": "Xuất ngoại trước kết hôn 2: Ngày bắt đầu", "結婚前渡航2 終了日": "Xuất ngoại trước kết hôn 2: Ngày kết thúc",
        "結婚前渡航3 開始日": "Xuất ngoại trước kết hôn 3: Ngày bắt đầu", "結婚前渡航3 終了日": "Xuất ngoại trước kết hôn 3: Ngày kết thúc",
        "結婚前渡航4 開始日": "Xuất ngoại trước kết hôn 4: Ngày bắt đầu", "結婚前渡航4 終了日": "Xuất ngoại trước kết hôn 4: Ngày kết thúc",
        "結婚前渡航5 開始日": "Xuất ngoại trước kết hôn 5: Ngày bắt đầu", "結婚前渡航5 終了日": "Xuất ngoại trước kết hôn 5: Ngày kết thúc",
        "結婚後の渡航回数": "Số lần xuất ngoại sau khi kết hôn",
        "結婚後渡航1 開始日": "Xuất ngoại sau kết hôn 1: Ngày bắt đầu", "結婚後渡航1 終了日": "Xuất ngoại sau kết hôn 1: Ngày kết thúc",
        "結婚後渡航2 開始日": "Xuất ngoại sau kết hôn 2: Ngày bắt đầu", "結婚後渡航2 終了日": "Xuất ngoại sau kết hôn 2: Ngày kết thúc",
        "結婚後渡航3 開始日": "Xuất ngoại sau kết hôn 3: Ngày bắt đầu", "結婚後渡航3 終了日": "Xuất ngoại sau kết hôn 3: Ngày kết thúc",
        "結婚後渡航4 開始日": "Xuất ngoại sau kết hôn 4: Ngày bắt đầu", "結婚後渡航4 終了日": "Xuất ngoại sau kết hôn 4: Ngày kết thúc",
        "結婚後渡航5 開始日": "Xuất ngoại sau kết hôn 5: Ngày bắt đầu", "結婚後渡航5 終了日": "Xuất ngoại sau kết hôn 5: Ngày kết thúc",
        # --- 退去強制歴 ---
        "申請人の退去強制歴の有無": "Tiền sử bị trục xuất của người nộp đơn",
        "退去強制された回数": "Số lần bị trục xuất",
        "違反内容:不法残留(オーバーステイ)": "Nội dung vi phạm: Cư trú bất hợp pháp (Overstay)",
        "違反内容:不法入国": "Nội dung vi phạm: Nhập cảnh trái phép",
        "違反内容:その他": "Nội dung vi phạm: Khác",
        "その他の具体内容": "Chi tiết vi phạm khác",
        "退去強制された年月日(直近)": "Ngày bị trục xuất (Gần nhất)",
        "出国した空港名": "Tên sân bay xuất cảnh",
        "当時のパスポート情報と今回の申請情報": "Thông tin hộ chiếu lúc đó và thông tin đăng ký lần này",
        "当時の国籍": "Quốc tịch lúc đó",
        "当時の氏名": "Họ tên lúc đó",
        "当時の生年月日": "Ngày sinh lúc đó",
        "退去強制までに同居した期間(開始)": "Thời gian sống chung trước khi bị trục xuất (Bắt đầu)",
        "退去強制までに同居した期間(終了)": "Thời gian sống chung trước khi bị trục xuất (Kết thúc)",
        "同居の住所": "Địa chỉ sống chung",
        # --- 親族情報 ---
        "夫の父:氏名": "Bố của chồng: Họ tên", "夫の父:年齢": "Bố của chồng: Tuổi", "夫の父:住所": "Bố của chồng: Địa chỉ", "夫の父:電話番号": "Bố của chồng: Số điện thoại",
        "夫の母:氏名": "Mẹ của chồng: Họ tên", "夫の母:年齢": "Mẹ của chồng: Tuổi", "夫の母:住所": "Mẹ của chồng: Địa chỉ", "夫の母:電話番号": "Mẹ của chồng: Số điện thoại",
        "夫の兄弟姉妹1:続柄": "Anh chị em của chồng 1: Quan hệ", "夫の兄弟姉妹1:氏名": "Anh chị em của chồng 1: Họ tên", "夫の兄弟姉妹1:年齢": "Anh chị em của chồng 1: Tuổi", "夫の兄弟姉妹1:住所": "Anh chị em của chồng 1: Địa chỉ", "夫の兄弟姉妹1:電話番号": "Anh chị em của chồng 1: Số điện thoại",
        "夫の兄弟姉妹2:続柄": "Anh chị em của chồng 2: Quan hệ", "夫の兄弟姉妹2:氏名": "Anh chị em của chồng 2: Họ tên", "夫の兄弟姉妹2:年齢": "Anh chị em của chồng 2: Tuổi", "夫の兄弟姉妹2:住所": "Anh chị em của chồng 2: Địa chỉ", "夫の兄弟姉妹2:電話番号": "Anh chị em của chồng 2: Số điện thoại",
        "夫の兄弟姉妹3:続柄": "Anh chị em của chồng 3: Quan hệ", "夫の兄弟姉妹3:氏名": "Anh chị em của chồng 3: Họ tên", "夫の兄弟姉妹3:年齢": "Anh chị em của chồng 3: Tuổi", "夫の兄弟姉妹3:住所": "Anh chị em của chồng 3: Địa chỉ", "夫の兄弟姉妹3:電話番号": "Anh chị em của chồng 3: Số điện thoại",
        "妻の父:氏名": "Bố của vợ: Họ tên", "妻の父:年齢": "Bố của vợ: Tuổi", "妻の父:住所": "Bố của vợ: Địa chỉ", "妻の父:電話番号": "Bố của vợ: Số điện thoại",
        "妻の母:氏名": "Mẹ của vợ: Họ tên", "妻の母:年齢": "Mẹ của vợ: Tuổi", "妻の母:住所": "Mẹ của vợ: Địa chỉ", "妻の母:電話番号": "Mẹ của vợ: Số điện thoại",
        "妻の兄弟姉妹1:続柄": "Anh chị em của vợ 1: Quan hệ", "妻の兄弟姉妹1:氏名": "Anh chị em của vợ 1: Họ tên", "妻の兄弟姉妹1:年齢": "Anh chị em của vợ 1: Tuổi", "妻の兄弟姉妹1:住所": "Anh chị em của vợ 1: Địa chỉ", "妻の兄弟姉妹1:電話番号": "Anh chị em của vợ 1: Số điện thoại",
        "妻の兄弟姉妹2:続柄": "Anh chị em của vợ 2: Quan hệ", "妻の兄弟姉妹2:氏名": "Anh chị em của vợ 2: Họ tên", "妻の兄弟姉妹2:年齢": "Anh chị em của vợ 2: Tuổi", "妻の兄弟姉妹2:住所": "Anh chị em của vợ 2: Địa chỉ", "妻の兄弟姉妹2:電話番号": "Anh chị em của vợ 2: Số điện thoại",
        "妻の兄弟姉妹3:続柄": "Anh chị em của vợ 3: Quan hệ", "妻の兄弟姉妹3:氏名": "Anh chị em của vợ 3: Họ tên", "妻の兄弟姉妹3:年齢": "Anh chị em của vợ 3: Tuổi", "妻の兄弟姉妹3:住所": "Anh chị em của vợ 3: Địa chỉ", "妻の兄弟姉妹3:電話番号": "Anh chị em của vợ 3: Số điện thoại",
        "お子さん1:続柄": "Con 1: Quan hệ", "お子さん1:氏名": "Con 1: Họ tên", "お子さん1:生年月日": "Con 1: Ngày sinh", "お子さん1:住所": "Con 1: Địa chỉ",
        "お子さん2:続柄": "Con 2: Quan hệ", "お子さん2:氏名": "Con 2: Họ tên", "お子さん2:生年月日": "Con 2: Ngày sinh", "お子さん2:住所": "Con 2: Địa chỉ",
        "お子さん3:続柄": "Con 3: Quan hệ", "お子さん3:氏名": "Con 3: Họ tên", "お子さん3:生年月日": "Con 3: Ngày sinh", "お子さん3:住所": "Con 3: Địa chỉ",
        "お子さん4:続柄": "Con 4: Quan hệ", "お子さん4:氏名": "Con 4: Họ tên", "お子さん4:生年月日": "Con 4: Ngày sinh", "お子さん4:住所": "Con 4: Địa chỉ",
        "お子さん5:続柄": "Con 5: Quan hệ", "お子さん5:氏名": "Con 5: Họ tên", "お子さん5:生年月日": "Con 5: Ngày sinh", "お子さん5:住所": "Con 5: Địa chỉ",
        "夫側で結婚を知っている:父": "Phía chồng biết về việc kết hôn: Bố", "夫側で結婚を知っている:母": "Phía chồng biết về việc kết hôn: Mẹ",
        "夫側で結婚を知っている:兄": "Phía chồng biết về việc kết hôn: Anh trai", "夫側で結婚を知っている:弟": "Phía chồng biết về việc kết hôn: Em trai",
        "夫側で結婚を知っている:姉": "Phía chồng biết về việc kết hôn: Chị gái", "夫側で結婚を知っている:妹": "Phía chồng biết về việc kết hôn: Em gái", "夫側で結婚を知っている:子": "Phía chồng biết về việc kết hôn: Con",
        "妻側で結婚を知っている:父": "Phía vợ biết về việc kết hôn: Bố", "妻側で結婚を知っている:母": "Phía vợ biết về việc kết hôn: Mẹ",
        "妻側で結婚を知っている:兄": "Phía vợ biết về việc kết hôn: Anh trai", "妻側で結婚を知っている:弟": "Phía vợ biết về việc kết hôn: Em trai",
        "妻側で結婚を知っている:姉": "Phía vợ biết về việc kết hôn: Chị gái", "妻側で結婚を知っている:妹": "Phía vợ biết về việc kết hôn: Em gái", "妻側で結婚を知っている:子": "Phía vợ biết về việc kết hôn: Con",
        "署名年": "Năm ký tên", "署名月": "Tháng ký tên", "署名日": "Ngày ký tên", "配偶者の署名": "Chữ ký của Vợ/Chồng",
        # --- 選択肢・プレースホルダー・ヘルプテキスト ---
        "男": "Nam", "女": "Nữ", "無": "Không", "有": "Có", "自己所有": "Sở hữu", "借家": "Thuê",
        "写真": "Ảnh", "電話": "Điện thoại", "対面": "Gặp mặt trực tiếp", "E-mail": "E-mail", "その他": "Khác",
        "難しい(通訳必要)": "Khó (Cần phiên dịch)", "筆談/あいさつ程度": "Chỉ viết tay / Chào hỏi", "日常会話程度は可能": "Có thể giao tiếp hàng ngày", "会話に支障なし": "Giao tiếp không gặp trở ngại",
        "初婚": "Kết hôn lần đầu", "再婚": "Tái hôn", "離婚": "Ly hôn", "死別": "Góa", "同じ": "Giống nhau", "別の氏名等": "Tên khác, v.v.",
        "例: ベトナム、フィリピン、中国": "VD: Việt Nam, Philippines, Trung Quốc", "例: NGUYEN THI HOA": "VD: NGUYEN THI HOA",
        "例: 山田 太郎": "VD: Yamada Taro", "例: ヤマダ タロウ": "VD: Yamada Taro", "例: 日本": "VD: Nhật Bản",
        "例: 東京都新宿区西新宿2-8-1": "VD: 2-8-1 Nishi-Shinjuku, Shinjuku-ku, Tokyo", "例: 03-1234-5678": "VD: 03-1234-5678", "例: 090-1234-5678": "VD: 090-1234-5678",
        "例: 85000": "VD: 85000", "例: 2LDK": "VD: 2LDK", "例: 株式会社サンプル商事": "VD: Công ty TNHH Sample", "例: 営業職": "VD: Nhân viên kinh doanh",
        "例: 東京都渋谷区のカフェ": "VD: Quán cà phê ở Shibuya, Tokyo", "例: 日本語、英語、中国語": "VD: Tiếng Nhật, Tiếng Anh, Tiếng Trung", "例: ベトナム語": "VD: Tiếng Việt",
        "例: 観光、仕事、留学": "VD: Du lịch, Công việc, Du học", "例: 兄、姉、弟、妹": "VD: Anh trai, Chị gái, Em trai, Em gái", "例: 夫の長男、妻の長女": "VD: Con trai trưởng của chồng, Con gái trưởng của vợ",
        "同居者ありを選んだ場合のみ入力": "Chỉ nhập nếu chọn 'Có' người sống cùng", "借家を選んだ場合のみ": "Chỉ nhập nếu chọn 'Thuê'",
        "年月日を示しながら詳しく記載": "Mô tả chi tiết, bao gồm ngày tháng năm", "結婚相談所による紹介の場合は会社名": "Tên công ty nếu được giới thiệu qua trung tâm môi giới", "死亡の場合は「死亡」と記載": "Ghi 'Đã mất' nếu áp dụng",
        # --- グループ名・サイドバー ---
        "①身分事項(基本情報)": "1. Thông tin cơ bản",
        "②結婚に至った経緯": "2. Quá trình dẫn đến kết hôn",
        "③-A 紹介者": "3-A. Người giới thiệu",
        "③-B 夫婦間の言語": "3-B. Ngôn ngữ giữa vợ chồng",
        "④-A 言語理解度": "4-A. Mức độ hiểu ngôn ngữ",
        "④-B 通訳": "4-B. Người phiên dịch",
        "④-C 結婚届証人(日本国内婚のみ)": "4-C. Người làm chứng kết hôn (Chỉ tại Nhật Bản)",
        "⑤-A 結婚式・披露宴": "5-A. Lễ cưới / Tiệc cưới",
        "⑤-B 結婚歴": "5-B. Lịch sử hôn nhân",
        "⑤-C 申請人の来日歴": "5-C. Lịch sử đến Nhật của người nộp đơn",
        "⑥-A 配偶者の渡航歴(結婚前)": "6-A. Lịch sử xuất ngoại của Vợ/Chồng (Trước kết hôn)",
        "⑥-B 配偶者の渡航歴(結婚後)": "6-B. Lịch sử xuất ngoại của Vợ/Chồng (Sau kết hôn)",
        "⑥-C 退去強制歴": "6-C. Tiền sử bị trục xuất",
        "⑦-A 退去強制後の同居": "7-A. Sống chung sau khi bị trục xuất",
        "⑦-B 夫の親族": "7-B. Họ hàng nhà chồng",
        "⑦-C 妻の親族": "7-C. Họ hàng nhà vợ",
        "⑧-A お子さん": "8-A. Con cái",
        "⑧-B 結婚を知っている親族": "8-B. Họ hàng biết về việc kết hôn",
        "⑧-C 署名": "8-C. Chữ ký",
        "紹介者がいる(お見合い結婚・結婚相談所等)": "Có người giới thiệu (Mai mối, Trung tâm, v.v.)",
        "結婚式・披露宴を行った": "Đã tổ chức Lễ cưới / Tiệc cưới",
        "申請人(お相手)が再婚": "Người nộp đơn tái hôn",
        "配偶者(あなたのクライアント)が再婚": "Vợ/Chồng (Khách hàng) tái hôn",
        "申請人の来日歴あり": "Người nộp đơn từng đến Nhật Bản",
        "配偶者の申請人母国訪問歴あり": "Vợ/Chồng từng đến quê hương của người nộp đơn",
        "退去強制歴あり": "Có tiền sử bị trục xuất",
        "お子さんあり": "Có con",
        "日本国内で婚姻届出した": "Đã đăng ký kết hôn tại Nhật Bản",
        "通訳が必要(言語が通じない)": "Cần người phiên dịch (Bất đồng ngôn ngữ)",
        "申請人が日本語を学習した経験あり": "Người nộp đơn từng học tiếng Nhật",
        "紹介者が外国人": "Người giới thiệu là người nước ngoài",
        # --- アプリUI（ヘッダー・ボタン・注意書き）---
        "配偶者ビザ申請 ヒアリングシート": "Đơn xin visa vợ/chồng — Phiếu phỏng vấn",
        "全{n}項目": "Tổng cộng {n} mục",
        "⚠️ 入力内容は自動保存されません": "⚠️ Nội dung nhập KHÔNG được tự động lưu",
        "作業を中断する場合は、下部の「💾 入力データをJSON保存」ボタンで必ずデータを保存してください。次回そのJSONを読み込めば続きから入力できます。":
            "Nếu cần tạm dừng, hãy chắc chắn lưu dữ liệu bằng nút “💾 Lưu dữ liệu nhập dưới dạng JSON” ở phía dưới. Lần sau tải tệp JSON đó lên để nhập tiếp.",
        "質問書の作成": "Tạo bản câu hỏi",
        "✨ 質問書Excelを作成する": "✨ Tạo bản câu hỏi (Excel)",
        "📥 質問書Excelをダウンロード": "📥 Tải xuống bản câu hỏi (Excel)",
        "💾 入力データをJSON保存": "💾 Lưu dữ liệu nhập dưới dạng JSON",
        "管理者ログイン": "Đăng nhập quản trị",
        "パスワード": "Mật khẩu",
        "ログイン": "Đăng nhập",
        "パスワードが違います": "Sai mật khẩu",
        "先生に送信する": "Gửi cho luật sư",
        "送信しています…": "Đang gửi…",
        "✅ 先生に送信しました。ご入力ありがとうございました。": "✅ Đã gửi cho luật sư. Cảm ơn bạn đã nhập thông tin!",
        "⬇️ 送信できなかった場合：データを保存して先生にメールで送ってください": "⬇️ Nếu gửi thất bại: hãy lưu dữ liệu và gửi email cho luật sư",
        "（氏名未入力）": "(chưa nhập tên)",
        "入力が終わったら、いちばん下の「📨 先生に送信する」ボタンを押してください。入力内容がそのまま先生に届きます。": "Khi hoàn tất, hãy nhấn nút 「📨 Gửi cho luật sư」 ở cuối trang. Nội dung sẽ được gửi trực tiếp đến luật sư.",
        "📊 入力済み項目を確認": "📊 Xem các mục đã nhập",
        "🔄 入力をリセット": "🔄 Đặt lại toàn bộ",
        "入力済みの内容": "Nội dung đã nhập",
        "✅ 質問書を作成しました。下のボタンからダウンロードできます。": "✅ Đã tạo bản câu hỏi. Tải xuống bằng nút bên dưới.",
        "🗑️ 入力内容をすべてリセットしました。": "🗑️ Đã đặt lại toàn bộ nội dung nhập.",
        "⚙️ ケース設定(行政書士用)": "⚙️ Cài đặt hồ sơ (dành cho hành chính thư sĩ)",
         "該当する項目にチェックを入れると、関連する質問が表示されます": "Đánh dấu các mục phù hợp; câu hỏi liên quan sẽ hiển thị.",
        "表示中の項目数": "Số mục đang hiển thị",
        "表示項目数": "Số mục hiển thị"
    }
}

# 翻訳を呼び出すためのおまじない関数
def get_text(text, lang):
    if lang == "ja":
        return text
    # 選んだ言語の辞書に翻訳があればそれを、無ければ元の日本語を返す
    return TRANSLATIONS.get(lang, {}).get(text, text)

# サイドバーに言語切り替えプルダウンを作成
lang_dict = {
    "ja": "🇯🇵 日本語", 
    "en": "🇺🇸 English", 
    "zh": "🇨🇳 中文", 
    "vi": "🇻🇳 Tiếng Việt"
}
selected_lang = st.sidebar.selectbox("Language / 言語", options=list(lang_dict.keys()), format_func=lambda x: lang_dict[x])

# ============================================
# マスタスキーマ読み込み
# ============================================
@st.cache_data
def load_schema():
    schema_path = Path(__file__).parent / "master_schema.json"
    with open(schema_path, encoding='utf-8') as f:
        return json.load(f)

schema = load_schema()

# ============================================
# セッション状態の初期化 & リセット処理
# ============================================
# form_data はウィジェット描画より前に必ず用意しておく
if "form_data" not in st.session_state:
    st.session_state.form_data = {}

# 「入力をリセット」ボタンが押されたら、ウィジェットが描画される前にここで全消去する。
# 各ウィジェットは key=field_id / key=flag_xxx を持つため、form_data だけ消しても
# 画面の入力欄は元に戻らない。→ ウィジェット用キー自体を削除する必要がある。
if st.session_state.get("_do_reset"):
    # 入力欄(各フィールド)のキーを削除
    for _f in schema["fields"]:
        st.session_state.pop(_f["field_id"], None)
    # ケースフラグ(サイドバーのチェック)のキーを削除
    for _flag in schema["case_flags"]:
        st.session_state.pop(f"flag_{_flag['flag_id']}", None)
    # 入力データ本体・作成済みExcelを削除
    st.session_state.form_data = {}
    st.session_state.pop("xlsx_bytes", None)
    # フラグ初期化フラグも消す → 直後の初期化処理で全ONに戻る
    st.session_state.pop("_flags_initialized", None)
    # リセット完了フラグを下げ、完了メッセージを次の描画で表示
    st.session_state["_do_reset"] = False
    st.session_state["_reset_done"] = True

# 差し込むコード
# ============================================
# サイドバーのケースフラグを「初回のみ」全OFFで初期化する
# --------------------------------------------
# ※重要: st.checkbox は key を持つと、value=True よりも
#   「セッションに保存済みの値」が優先される。
#   ここで session_state に直接 False を書き込むことで、
#   キャッシュに負けず確実に全OFF（全て外れている状態）からスタートさせる。
if not st.session_state.get("_flags_initialized"):
    for _flag in schema["case_flags"]:
        st.session_state[f"flag_{_flag['flag_id']}"] = False
    st.session_state["_flags_initialized"] = True

# ============================================
# 転記エンジン読み込み
# ============================================
import sys
sys.path.insert(0, str(Path(__file__).parent))
import fill_questionnaire as fq

# 入管公式書式テンプレート(アプリと同じフォルダに置く)
TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"

# ============================================
# 管理者(先生)モード ＆ クライアント→先生 メール送信
# ============================================
import smtplib
from email.message import EmailMessage

if "is_admin" not in st.session_state:
    st.session_state["is_admin"] = False

# ============================================
# secrets.toml を堅牢に読む
#   1) st.secrets (Streamlit Cloud / .streamlit/secrets.toml)
#   2) アプリと同じフォルダの secrets.toml（フォールバック）
#   3) 環境変数 ADMIN_PASSWORD
# どれか1つでも値があればOK
# ============================================
import os

_LOCAL_SECRETS_CACHE = None

def _load_local_secrets():
    """アプリと同じフォルダ／.streamlit フォルダの secrets.toml を読む
    BOM付き / Shift-JIS でも落ちないようにする
    """
    global _LOCAL_SECRETS_CACHE
    if _LOCAL_SECRETS_CACHE is not None:
        return _LOCAL_SECRETS_CACHE
    _LOCAL_SECRETS_CACHE = {}
    here = Path(__file__).parent
    candidates = [
        here / ".streamlit" / "secrets.toml",
        here / "secrets.toml",
        Path.cwd() / ".streamlit" / "secrets.toml",
        Path.cwd() / "secrets.toml",
    ]
    try:
        import tomllib  # Python 3.11+
        _toml_loads = lambda s: tomllib.loads(s)
    except Exception:
        try:
            import toml as _toml
            _toml_loads = lambda s: _toml.loads(s)
        except Exception:
            _toml_loads = None
    for p in candidates:
        if not p.exists() or _toml_loads is None:
            continue
        for enc in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
            try:
                text = p.read_text(encoding=enc)
                _LOCAL_SECRETS_CACHE = _toml_loads(text) or {}
                _LOCAL_SECRETS_CACHE["__path__"] = str(p)
                _LOCAL_SECRETS_CACHE["__encoding__"] = enc
                return _LOCAL_SECRETS_CACHE
            except UnicodeDecodeError:
                continue
            except Exception:
                # TOMLパース失敗（クォート漏れ等）はその旨だけ記録
                _LOCAL_SECRETS_CACHE = {"__path__": str(p), "__error__": "TOML parse error"}
                return _LOCAL_SECRETS_CACHE
    return _LOCAL_SECRETS_CACHE


def _secret(path, default=None):
    """st.secrets → ローカル secrets.toml → 環境変数 の順で安全に値を取る。
    どれも無ければ default を返す（決して例外を投げない）。
    """
    # 1) st.secrets (Streamlit Cloud / .streamlit/secrets.toml が読めている場合)
    try:
        cur = st.secrets
        for p in path:
            cur = cur[p]
        if cur not in (None, ""):
            return cur
    except Exception:
        pass
    # 2) アプリと同じフォルダの secrets.toml をフォールバック
    try:
        cur = _load_local_secrets()
        for p in path:
            if isinstance(cur, dict):
                cur = cur.get(p)
            else:
                cur = None
                break
        if cur not in (None, ""):
            return cur
    except Exception:
        pass
    # 3) 環境変数（admin_password だけサポート）
    if path == ["admin_password"]:
        env = os.environ.get("ADMIN_PASSWORD")
        if env:
            return env
    return default


def _st_secrets_dict():
    """st.secrets を普通の dict として安全に取得（読めなければ空dict）"""
    try:
        return {k: v for k, v in st.secrets.items()}
    except Exception:
        return {}


def _secrets_shape():
    """設定の「キー名だけ」を一覧にする（値は絶対に表示しない）。診断用。"""
    shape = {}
    for src_name, src in (("st.secrets", _st_secrets_dict()), ("ローカルsecrets.toml", _load_local_secrets())):
        if not isinstance(src, dict) or not src:
            continue
        inner = {}
        for k, v in src.items():
            if k.startswith("__"):
                continue
            inner[k] = sorted(v.keys()) if isinstance(v, dict) or hasattr(v, "keys") else "（値あり）"
        if inner:
            shape[src_name] = inner
    return shape


def _find_admin_password():
    """admin_password を探す。戻り値: (パスワード, 見つかった場所の説明)
    正しい場所（トップレベル）→ セクション内に紛れ込んだ場合（例: [smtp] の下）
    → 環境変数 の順で探し、どこにあってもログインできるようにする。
    """
    v = _secret(["admin_password"])
    if v:
        src = "環境変数 ADMIN_PASSWORD" if (not _st_secrets_dict().get("admin_password")
              and not (_load_local_secrets() or {}).get("admin_password")
              and os.environ.get("ADMIN_PASSWORD")) else "トップレベル（正しい場所）"
        return str(v), src
    for src_name, src in (("st.secrets", _st_secrets_dict()), ("ローカルsecrets.toml", _load_local_secrets())):
        if not isinstance(src, dict):
            continue
        for sec, val in src.items():
            if sec.startswith("__"):
                continue
            try:
                if (isinstance(val, dict) or hasattr(val, "get")) and val.get("admin_password"):
                    return str(val["admin_password"]), f"{src_name} の [{sec}] セクション内（※本来はトップレベル推奨）"
            except Exception:
                continue
    return "", ""


def send_to_sensei(json_text, client_name):
    """入力データ(JSON)を先生へメール送信する。戻り値: (成功フラグ, エラーメッセージ)"""
    def _smtp(key, default=None):
        # [smtp] セクション → smtp_xxx → トップレベル の順で探す（貼り間違い救済）
        v = _secret(["smtp", key])
        if v in (None, ""):
            v = _secret([f"smtp_{key}"])
        if v in (None, ""):
            v = _secret([key], default)
        return v

    host = _smtp("host")
    user = _smtp("user")
    password = _smtp("password")
    port = int(_smtp("port", 465) or 465)
    to_addr = _smtp("to_email", user) or user
    if not (host and user and password):
        return False, "メール設定が未登録のため送信できません。お手数ですが、先生に直接ご連絡ください。"

    msg = EmailMessage()
    msg["Subject"] = "配偶者ビザヒアリングシート送信"
    msg["From"] = user
    msg["To"] = to_addr
    msg.set_content(
        f"送信日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"クライアント名: {client_name}\n\n"
        "添付のJSONファイルを、質問書アプリの管理者モード（サイドバー）で読み込んでください。"
    )
    msg.add_attachment(
        json_text.encode("utf-8"),
        maintype="application", subtype="json",
        filename=f"hearing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
    )
    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port, timeout=30) as s:
                s.login(user, password)
                s.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=30) as s:
                s.ehlo()
                s.starttls()
                s.login(user, password)
                s.send_message(msg)
        return True, ""
    except Exception as e:
        return False, f"送信できませんでした: {e}"

# ============================================
# 体験版モード設定
#   DEMO_MODE = True  → シート1(1ページ目)のみ生成する無料体験版
#   DEMO_MODE = False → 全8ページを生成する製品版(購入者向け)
# ============================================
DEMO_MODE = False
COCONALA_URL = "https://coconala.com/services/4240552"   # ←ここにあなたのココナラ商品URLを貼ってください


def _demo_sheet1_only(full_bytes):
    """全8ページのExcelから「シート1」だけを残し、体験版の透かしを付けて返す。"""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(full_bytes))
    # シート1以外を削除（"Sheet1" を残す）
    for ws in list(wb.worksheets):
        if ws.title != "Sheet1":
            wb.remove(ws)
    ws1 = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    # 印刷/PDFのヘッダーに体験版の透かしを入れる
    ws1.oddHeader.center.text = "＜無料体験版 SAMPLE＞  全8ページ版は TransLayer / ココナラ にて"
    ws1.oddHeader.center.size = 9
    ws1.oddHeader.center.font = "ＭＳ 明朝"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_data():
    return {
        "case_flags": case_flags,
        "form_data": st.session_state.form_data,
        "saved_at": datetime.now().isoformat(),
    }


# ============================================
# ヘッダー
# ============================================
st.title("💞 " + get_text("配偶者ビザ申請 ヒアリングシート", selected_lang))
_total = schema['meta']['statistics']['total_fields']
st.caption(f"📋 {get_text(schema['meta']['document_name'], selected_lang)} v{schema['meta']['version']}  | "
           + get_text("全{n}項目", selected_lang).format(n=_total))

# 体験版バナー
if DEMO_MODE:
    st.info(
        "🎁 **これは無料体験版です。** 入力すると **シート1（1ページ目）だけ** をExcelで生成できます。"
        "出力の品質・レイアウトをそのままお試しください。\n\n"
        f"➡️ **全8ページ版**(質問書まるごと)は [TransLayer / ココナラ]({COCONALA_URL}) で販売中です。"
    )

# 入力内容は保存されない旨の注意書き（途中で中断する人向け）
if st.session_state["is_admin"]:
    st.warning(
        "**" + get_text("⚠️ 入力内容は自動保存されません", selected_lang) + "**\n\n"
        + get_text(
            "作業を中断する場合は、下部の「💾 入力データをJSON保存」ボタンで必ずデータを保存してください。"
            "次回そのJSONを読み込めば続きから入力できます。",
            selected_lang,
        )
    )
else:
    st.warning(
        "**" + get_text("⚠️ 入力内容は自動保存されません", selected_lang) + "**\n\n"
        + get_text(
            "入力が終わったら、いちばん下の「📨 先生に送信する」ボタンを押してください。入力内容がそのまま先生に届きます。",
            selected_lang,
        )
    )

# ============================================
# サイドバー: 管理者(先生)ログイン
# ============================================
with st.sidebar:
    if not st.session_state["is_admin"]:
        with st.expander("🔑 " + get_text("管理者ログイン", selected_lang)):
            _pw = st.text_input(get_text("パスワード", selected_lang), type="password", key="_admin_pw")
            if st.button(get_text("ログイン", selected_lang), key="_admin_login", use_container_width=True):
                _correct, _pw_src = _find_admin_password()
                if not _correct:
                    # secrets が一切読めていない場合は具体的な原因を案内する
                    _local = _load_local_secrets()
                    _msg_lines = [
                        "⚠️ パスワード設定（admin_password）が読み込めていません。",
                        "",
                        "次のいずれかを設定してください：",
                        "① `.streamlit/secrets.toml` をアプリと同じフォルダに置く",
                        "② 環境変数 `ADMIN_PASSWORD` を設定する",
                        "③ Streamlit Cloud の場合：Settings → Secrets の貼り付け内容で、`admin_password = \"...\"` の行が `[smtp]` よりも**上**にあるか確認する",
                    ]
                    if _local.get("__error__"):
                        _msg_lines.append(f"❌ secrets.toml の書式エラー: {_local['__path__']}")
                    elif _local.get("__path__"):
                        _msg_lines.append(f"📄 ファイルは見つかりましたが admin_password がありません: {_local['__path__']}")
                    _shape = _secrets_shape()
                    if _shape:
                        _msg_lines.append("")
                        _msg_lines.append("🔎 いま読めている設定キー（値は表示しません）:")
                        _msg_lines.append(f"`{json.dumps(_shape, ensure_ascii=False)}`")
                    st.error("\n\n".join(_msg_lines))
                elif _pw == _correct:
                    st.session_state["is_admin"] = True
                    st.rerun()
                else:
                    st.error(get_text("パスワードが違います", selected_lang))
    else:
        st.success("🔓 管理者モード（先生用）")
        _up = st.file_uploader("📂 クライアントから届いたJSONを読み込む", type=["json"], key="_admin_json_up")
        if _up is not None:
            _sig = f"{_up.name}:{_up.size}"
            if st.session_state.get("_loaded_json_sig") != _sig:
                try:
                    _loaded = json.loads(_up.getvalue().decode("utf-8"))
                    _fd = dict(_loaded.get("form_data") or {})
                    # 他言語で入力された選択肢 → 日本語へ逆変換するための辞書
                    _rev = {}
                    for _ld in TRANSLATIONS.values():
                        for _jp, _tr in _ld.items():
                            _rev.setdefault(_tr, _jp)
                    for _f in schema["fields"]:
                        _fid2 = _f["field_id"]
                        st.session_state.pop(_fid2, None)  # ウィジェットを作り直させる(リセットと同方式)
                        if _fid2 not in _fd:
                            continue
                        _v = _fd[_fid2]
                        _w = _f.get("widget", "text_input")
                        if _w == "date_input" and isinstance(_v, str) and _v:
                            try:
                                _fd[_fid2] = date.fromisoformat(_v[:10])
                            except Exception:
                                pass
                        elif _w in ("selectbox", "radio") and isinstance(_v, str):
                            _fd[_fid2] = _rev.get(_v, _v)
                    st.session_state.form_data = _fd
                    for _fk, _fv in (_loaded.get("case_flags") or {}).items():
                        st.session_state[f"flag_{_fk}"] = bool(_fv)
                    st.session_state.pop("xlsx_bytes", None)
                    st.session_state["_loaded_json_sig"] = _sig
                    st.rerun()
                except Exception as e:
                    st.error(f"JSONを読み込めませんでした: {e}")
            else:
                st.caption("✅ 読み込み済み: " + _up.name)
        if st.button("ログアウト", key="_admin_logout", use_container_width=True):
            st.session_state["is_admin"] = False
            st.rerun()
    st.markdown("---")

# ============================================
# サイドバー: 行政書士用ケースフラグ設定
# ============================================
st.sidebar.header(get_text("⚙️ ケース設定(行政書士用)", selected_lang))
st.sidebar.caption(get_text("該当する項目にチェックを入れると、関連する質問が表示されます", selected_lang))

case_flags = {}
for flag in schema["case_flags"]:
    # 初期値は上の初期化処理で session_state に全ON で書き込み済み。
    # ここでは value を渡さず、session_state の値をそのまま使う。
    # 親フラグに依存するフラグは、親がONの時だけ表示
    if flag.get("depends_on"):
        parent_value = case_flags.get(flag["depends_on"], False)
        if not parent_value:
            case_flags[flag["flag_id"]] = False
            continue
    case_flags[flag["flag_id"]] = st.sidebar.checkbox(
        get_text(flag["label"], selected_lang),
        key=f"flag_{flag['flag_id']}"
    )

st.sidebar.markdown("---")
st.sidebar.subheader("📊 " + get_text("表示中の項目数", selected_lang))

# フィールドの表示判定
def is_field_visible(field, flags):
    """このフィールドを表示するか判定"""
    if field["field_type"] == "common" or field["field_type"] == "required":
        return True
    # differential の場合、依存フラグを確認
    dep_flag = field.get("depends_on_flag")
    if dep_flag is None:
        return True  # フラグ指定なしの差分項目は常に表示
    return flags.get(dep_flag, False)

visible_fields = [f for f in schema["fields"] if is_field_visible(f, case_flags)]
st.sidebar.metric(get_text("表示項目数", selected_lang), f"{len(visible_fields)} / {len(schema['fields'])}")

# ============================================
# メインエリア: グループごとにフォーム表示
# ============================================
# グループでフィールドを束ねる
groups = defaultdict(list)
for fld in visible_fields:
    groups[fld.get("group", "その他")].append(fld)

# ウィジェット描画関数
def render_widget(field):
    fid = field["field_id"]
    label = get_text(field["label"], selected_lang) + (" 🔴" if field.get("required") else "")
    help_text = get_text(field.get("help_text"), selected_lang)
    placeholder = get_text(field.get("placeholder", ""), selected_lang)
    widget = field.get("widget", "text_input")
    options = [get_text(opt, selected_lang) for opt in field.get("options", [])]

    # 既存値の取得
    current = st.session_state.form_data.get(fid)
    
    if widget == "text_input":
        val = st.text_input(label, value=current or "", 
                            placeholder=placeholder, help=help_text, key=fid)
    elif widget == "text_area":
        val = st.text_area(label, value=current or "",
                           placeholder=placeholder, help=help_text, 
                           max_chars=field.get("max_length"), key=fid, height=120)
    elif widget == "date_input":
        val = st.date_input(
            label, 
            value=current if isinstance(current, date) else None,
            min_value=date(1920, 1, 1),  # ここを追加！1920年まで遡れるようにします
            max_value=date.today(),      # ここを追加！未来の日付を選べなくします
            help=help_text, 
            key=fid, 
            format="YYYY-MM-DD"
        )
        val = val.isoformat() if val else ""
    elif widget == "number_input":
        val = st.number_input(label, value=int(current) if current else 0,
                              help=help_text, key=fid, step=1)
    elif widget == "selectbox":
        idx = options.index(current) if current in options else 0
        val = st.selectbox(label, options, index=idx, help=help_text, key=fid)
    elif widget == "radio":
        idx = options.index(current) if current in options else 0
        val = st.radio(label, options, index=idx, help=help_text, key=fid,
                       horizontal=True)
    elif widget == "checkbox":
        val = st.checkbox(label, value=bool(current), help=help_text, key=fid)
    else:
        val = st.text_input(label, value=current or "", key=fid)
    
    st.session_state.form_data[fid] = val
    return val

# タブで8ページを切り替え
sheet_groups = defaultdict(list)
for group_name, flds in groups.items():
    # グループに属する1つ目のフィールドのシート名でタブ分け
    sheet = flds[0]["sheet"] if flds else "Sheet1"
    sheet_groups[sheet].append((group_name, flds))

sheets_in_order = sorted(sheet_groups.keys(), key=lambda x: x.replace("Sheet", "").replace("４", "4"))
tabs = st.tabs([f"📄 {s}" for s in sheets_in_order])

for tab, sheet_name in zip(tabs, sheets_in_order):
    with tab:
      for group_name, flds in sheet_groups[sheet_name]:
            with st.expander(f"📌 {get_text(group_name, selected_lang)}", expanded=True):

                # 2カラムで配置
                cols = st.columns(2)
                for i, fld in enumerate(flds):
                    with cols[i % 2]:
                        render_widget(fld)

# ============================================
# フッター: 質問書作成 & 保存ボタン
# ============================================
st.markdown("---")
st.subheader("📄 " + get_text("質問書の作成", selected_lang))

# リセット完了メッセージ（リセット直後の再描画で表示）
if st.session_state.pop("_reset_done", False):
    st.success(get_text("🗑️ 入力内容をすべてリセットしました。", selected_lang))

cA, cB = st.columns([2, 1])

with cA:
    if st.session_state["is_admin"]:
        # ---------- 管理者(先生)モード: Excel作成 ----------
        if st.button(get_text("✨ 質問書Excelを作成する", selected_lang), type="primary", use_container_width=True):
            if not TEMPLATE_PATH.exists():
                st.error(f"テンプレートが見つかりません: {TEMPLATE_PATH.name} をアプリと同じフォルダに置いてください。")
            else:
                try:
                    xlsx_bytes = fq.fill_to_bytes(str(TEMPLATE_PATH), build_data())
                    if DEMO_MODE:
                        xlsx_bytes = _demo_sheet1_only(xlsx_bytes)   # シート1のみに絞る
                    st.session_state["xlsx_bytes"] = xlsx_bytes
                    if DEMO_MODE:
                        st.success("✅ 体験版（シート1のみ）を作成しました。下のボタンからダウンロードできます。")
                        st.warning(
                            "これは体験版のため **1ページ目のみ** です。"
                            f"全8ページの完成版は [TransLayer / ココナラ]({COCONALA_URL}) からご利用いただけます。"
                        )
                    else:
                        st.success(get_text("✅ 質問書を作成しました。下のボタンからダウンロードできます。", selected_lang))
                except Exception as e:
                    st.error(f"作成中にエラーが発生しました: {e}")

        if st.session_state.get("xlsx_bytes"):
            applicant = st.session_state.form_data.get("applicant_name", "申請人")
            safe = "".join(ch for ch in str(applicant) if ch.isalnum() or ch in " _-")[:20].strip() or "申請人"
            _label = "📥 体験版Excel（シート1）をダウンロード" if DEMO_MODE else get_text("📥 質問書Excelをダウンロード", selected_lang)
            _prefix = "質問書_体験版SAMPLE" if DEMO_MODE else "質問書"
            st.download_button(
                _label,
                data=st.session_state["xlsx_bytes"],
                file_name=f"{_prefix}_{safe}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    else:
        # ---------- クライアントモード: 先生に送信 ----------
        if st.button("📨 " + get_text("先生に送信する", selected_lang), type="primary", use_container_width=True):
            _client_name = str(st.session_state.form_data.get("applicant_name") or "").strip() or get_text("（氏名未入力）", selected_lang)
            _json_text = json.dumps(build_data(), ensure_ascii=False, indent=2, default=str)
            with st.spinner(get_text("送信しています…", selected_lang)):
                _ok, _err = send_to_sensei(_json_text, _client_name)
            st.session_state["_sent_ok"] = _ok
            st.session_state["_sent_err"] = _err
            st.session_state["_fallback_json"] = _json_text
        if st.session_state.get("_sent_ok") is True:
            st.success(get_text("✅ 先生に送信しました。ご入力ありがとうございました。", selected_lang))
        elif st.session_state.get("_sent_ok") is False:
            st.error(st.session_state.get("_sent_err") or "")
            st.download_button(
                get_text("⬇️ 送信できなかった場合：データを保存して先生にメールで送ってください", selected_lang),
                data=st.session_state.get("_fallback_json", "{}"),
                file_name=f"hearing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                use_container_width=True,
            )

with cB:
    filled = {k: v for k, v in st.session_state.form_data.items() if v}
    st.metric(get_text("入力済み項目", selected_lang), f"{len(filled)} / {len(visible_fields)}")

st.markdown("---")
c1, c2, c3 = st.columns([1, 1, 1])

with c1:
    if st.session_state["is_admin"]:
        out = build_data()
        st.download_button(
            get_text("💾 入力データをJSON保存", selected_lang),
            data=json.dumps(out, ensure_ascii=False, indent=2, default=str),
            file_name=f"hearing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json",
            use_container_width=True,
        )

with c2:
    if st.button(get_text("📊 入力済み項目を確認", selected_lang), use_container_width=True):
        with st.expander(get_text("入力済みの内容", selected_lang), expanded=True):
            st.json({k: v for k, v in st.session_state.form_data.items() if v})

with c3:
    # リセットは「フラグを立てて再実行」方式。
    # 実際の消去はスクリプト冒頭(ウィジェット描画前)で行うため、入力欄が確実に空に戻る。
    if st.button(get_text("🔄 入力をリセット", selected_lang), use_container_width=True):
        st.session_state["_do_reset"] = True
        st.rerun()

# ============================================
# デバッグエリア(開発時のみ)
# ============================================
with st.sidebar.expander("🔧 デバッグ情報"):
    # secrets が読めているかの診断（パスワード設定で詰まる人向け）
    st.write("**🔐 secrets 診断:**")
    _pw_val, _pw_src = _find_admin_password()
    _local = _load_local_secrets()
    st.write({
        "admin_password": ("✅ あり（" + _pw_src + "）") if _pw_val else "❌ 見つかりません",
        "ローカル secrets.toml のパス": _local.get("__path__", "（見つかりません）"),
        "ローカル secrets.toml の文字コード": _local.get("__encoding__", "-"),
        "ローカル secrets.toml のエラー": _local.get("__error__", "-"),
        "環境変数 ADMIN_PASSWORD": "設定あり" if os.environ.get("ADMIN_PASSWORD") else "未設定",
        "メール送信 host/user/password": "/".join(
            ("✅" if _secret(["smtp", k]) or _secret([f"smtp_{k}"]) or _secret([k]) else "❌")
            for k in ("host", "user", "password")
        ),
    })
    _shape = _secrets_shape()
    if _shape:
        st.write("**🔎 読めている設定キー（値は出しません）:**")
        st.json(_shape)
    st.write("**ケースフラグ:**")
    st.json(case_flags)
